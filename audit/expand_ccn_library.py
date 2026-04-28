"""
Layer 5 step — repopulate the CCN_Library sheet of
BFR_Generator_FC2-000-05N.xlsx with all 1,059 CCNs from
audit/CCN_VOCABULARY.json (extracted from FC 2-000-05N Appendix A,
2019-06-27).

Preserves the 23 curated entries (which carry hand-authored Factor /
Factor Notes / NTG / Driver Description values from the original
methodology workbook author) by merging on CCN. New entries inherit
title from the canonical vocabulary; factor/NTG/notes left blank for
SME population per project.

CCN format harmonization: the existing CCN_Library uses 5-char
spaced format `"143 70"`; the canonical vocabulary uses 5-digit
unspaced `"14170"`. All emitted CCN values are written in spaced form
so existing BFR_Calculator entries (which use spaced format) still
resolve via VLOOKUP.

CCN_TABLE named range is expanded from CCN_Library!$C$6:$I$28 to
CCN_Library!$C$6:$I$1100 (covers the merged list with headroom for
future additions).

Run:  python3 audit/expand_ccn_library.py
"""

import json
import re
from pathlib import Path

import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CCN_PATTERN = re.compile(r"^\d{3}\s?\d{2}$")  # "143 70" or "14370"

REPO_ROOT = Path(__file__).resolve().parent.parent
WB_PATH = REPO_ROOT / "BFR_Generator_FC2-000-05N.xlsx"
VOCAB_PATH = REPO_ROOT / "audit" / "CCN_VOCABULARY.json"

HEADER_ROW = 5
DATA_START_ROW = 6
CCN_TABLE_RIGHT_BUFFER_ROW = 1100


def to_spaced(ccn_5digit: str) -> str:
    """Convert canonical '14170' -> spaced '141 70' for sheet display."""
    if ccn_5digit and len(ccn_5digit) == 5 and ccn_5digit.isdigit():
        return f"{ccn_5digit[:3]} {ccn_5digit[3:]}"
    return ccn_5digit


def to_unspaced(ccn_spaced: str) -> str:
    """Inverse: '141 70' -> '14170'."""
    if not ccn_spaced:
        return ccn_spaced
    return str(ccn_spaced).replace(" ", "")


def vocab_uom(rec):
    """Pick a printable UoM for the library row.

    Source vocabulary has uom_area / uom_other / uom_alt as bracketed
    or plain tokens (e.g. '[SF]', '[SY]', 'LF', 'AC'). For the library's
    UM column, use the cleanest non-bracket form of uom_area when
    present; fall back to other / alt.
    """
    for key in ("uom_area", "uom_other", "uom_alt"):
        v = rec.get(key)
        if v:
            return v.strip("[]")
    return None


def main():
    wb = openpyxl.load_workbook(WB_PATH)
    lib = wb["CCN_Library"]

    # 1. Snapshot existing curated entries — only rows whose column C
    #    matches a CCN pattern (3-digit + optional space + 2-digit).
    #    The original sheet had non-CCN data (e.g., Vehicle Sub-Schedule
    #    Type A..G rows) further down that must NOT enter the dictionary.
    curated = {}
    for r in range(DATA_START_ROW, lib.max_row + 1):
        ccn_cell = lib.cell(row=r, column=3).value
        if not ccn_cell:
            continue
        ccn_str = str(ccn_cell).strip()
        if not CCN_PATTERN.match(ccn_str):
            continue
        key = to_unspaced(ccn_str)
        curated[key] = {
            "ccn_display": ccn_cell,
            "facility_name": lib.cell(row=r, column=4).value,
            "uom": lib.cell(row=r, column=5).value,
            "default_factor": lib.cell(row=r, column=6).value,
            "factor_notes": lib.cell(row=r, column=7).value,
            "ntg": lib.cell(row=r, column=8).value,
            "driver_desc": lib.cell(row=r, column=9).value,
        }
    print(f"Captured {len(curated)} curated entries from existing CCN_Library.")

    # 2. Load canonical vocabulary
    vocab = json.loads(VOCAB_PATH.read_text())
    print(f"Loaded {vocab['ccn_count']} CCNs from {VOCAB_PATH.name} "
          f"(source: {vocab['provenance']['document_title']}, "
          f"{vocab['provenance']['source_report_date']}).")

    # 3. Build merged set keyed by 5-digit CCN
    merged = {}
    for v in vocab["ccns"]:
        key = v["ccn"]
        merged[key] = {
            "ccn_display": to_spaced(key),
            "facility_name": v.get("title") or "",
            "uom": vocab_uom(v),
            "default_factor": None,
            "factor_notes": None,
            "ntg": None,
            "driver_desc": v.get("description") or "",
            "source_page": v.get("page"),
        }
    # Overlay curated values (curated wins on factor / notes / ntg /
    # driver_desc / facility_name when present)
    for key, c in curated.items():
        if key in merged:
            for col in ("facility_name", "uom", "default_factor",
                        "factor_notes", "ntg", "driver_desc"):
                if c.get(col) is not None and c.get(col) != "":
                    merged[key][col] = c[col]
        else:
            merged[key] = {
                "ccn_display": c["ccn_display"],
                "facility_name": c["facility_name"],
                "uom": c["uom"],
                "default_factor": c["default_factor"],
                "factor_notes": c["factor_notes"],
                "ntg": c["ntg"],
                "driver_desc": c["driver_desc"],
                "source_page": None,
            }

    sorted_keys = sorted(merged.keys())
    print(f"Merged set: {len(sorted_keys)} CCNs "
          f"({len(curated)} curated overlaid on {vocab['ccn_count']} canonical).")

    # 4. Unmerge any merged cells in the data region, then clear values
    merged_to_remove = [
        rng for rng in list(lib.merged_cells.ranges)
        if rng.min_row >= DATA_START_ROW
    ]
    for rng in merged_to_remove:
        lib.unmerge_cells(str(rng))
    for r in range(DATA_START_ROW, lib.max_row + 1):
        for c in range(1, lib.max_column + 1):
            lib.cell(row=r, column=c).value = None

    # 5. Write merged data
    INPUT_FILL = PatternFill("solid", fgColor="FFF8DC")  # cream
    CALC_FILL = PatternFill("solid", fgColor="EAF3F4")   # light teal
    thin = Side(border_style="thin", color="000000")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    BODY_FONT = Font(name="Calibri", size=10)

    for i, key in enumerate(sorted_keys):
        r = DATA_START_ROW + i
        rec = merged[key]
        cells = [
            (3, rec["ccn_display"]),
            (4, rec["facility_name"]),
            (5, rec["uom"]),
            (6, rec["default_factor"]),
            (7, rec["factor_notes"]),
            (8, rec["ntg"]),
            (9, rec["driver_desc"]),
        ]
        for col, val in cells:
            cell = lib.cell(row=r, column=col, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True)

    last_data_row = DATA_START_ROW + len(sorted_keys) - 1
    print(f"Wrote rows {DATA_START_ROW}..{last_data_row}.")

    # 6. Update CCN_TABLE named range (delete + re-add)
    if "CCN_TABLE" in wb.defined_names:
        del wb.defined_names["CCN_TABLE"]
    new_range = f"CCN_Library!$C${DATA_START_ROW}:$I${CCN_TABLE_RIGHT_BUFFER_ROW}"
    wb.defined_names["CCN_TABLE"] = DefinedName(
        name="CCN_TABLE", attr_text=new_range
    )
    print(f"CCN_TABLE -> {new_range}")

    # 7. Preserve full-recalc-on-open flag
    wb.calculation.fullCalcOnLoad = True

    wb.save(WB_PATH)
    print(f"Saved: {WB_PATH}")


if __name__ == "__main__":
    main()
