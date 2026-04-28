"""Structural inventory of an Excel workbook.

Emits a text report describing:
  - sheet list with dimensions and visibility
  - defined names (named ranges)
  - external links / references
  - per-sheet: merged cells count, frozen panes, tab color, print area
  - presence of formulas, data validations, conditional formats
  - rows that look like CCN-bearing rows (heuristic: token like NNN-NN, NNN NN, etc.)
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

CCN_RE = re.compile(r"\b\d{3}[\s\-\.]?\d{2}\b")


def inventory(path: Path) -> str:
    out: list[str] = []
    p = path.resolve()
    out.append(f"FILE: {p.name}")
    out.append(f"PATH: {p}")
    out.append(f"SIZE: {p.stat().st_size:,} bytes")
    out.append("")

    # data_only=False to preserve formulas; we'll do a second pass for cached values.
    wb = load_workbook(filename=str(p), data_only=False, read_only=False, keep_links=True)

    out.append(f"SHEETS ({len(wb.sheetnames)}):")
    for i, name in enumerate(wb.sheetnames):
        ws = wb[name]
        out.append(
            f"  [{i:02d}] {name!r}  state={ws.sheet_state}  "
            f"dim={ws.dimensions}  max_row={ws.max_row}  max_col={ws.max_column}"
        )
    out.append("")

    # Defined names
    dn = list(wb.defined_names)
    out.append(f"DEFINED NAMES ({len(dn)}):")
    for nm in dn:
        d = wb.defined_names[nm]
        out.append(f"  {nm!r}  ->  {d.value!r}")
    out.append("")

    # External links
    try:
        ext = wb._external_links  # noqa: SLF001
    except Exception:
        ext = []
    out.append(f"EXTERNAL LINKS ({len(ext)}):")
    for link in ext:
        target = getattr(getattr(link, "file_link", None), "Target", None)
        out.append(f"  target={target!r}")
    out.append("")

    # Per-sheet detail
    out.append("PER-SHEET DETAIL")
    out.append("=" * 60)
    for name in wb.sheetnames:
        ws = wb[name]
        out.append(f"\n[Sheet: {name!r}]")
        out.append(f"  state         : {ws.sheet_state}")
        out.append(f"  tab_color     : {getattr(ws.sheet_properties.tabColor, 'value', None)}")
        out.append(f"  freeze_panes  : {ws.freeze_panes}")
        out.append(f"  merged_cells  : {len(ws.merged_cells.ranges)}")
        out.append(f"  print_area    : {ws.print_area}")
        out.append(f"  data_validns  : {len(ws.data_validations.dataValidation)}")
        out.append(f"  cond_formats  : {sum(len(v) for v in ws.conditional_formatting._cf_rules.values())}")  # noqa: SLF001

        # Scan formulas / CCN tokens. Cap scan to keep memory bounded.
        max_r = min(ws.max_row or 0, 4000)
        max_c = min(ws.max_column or 0, 80)
        formula_count = 0
        ext_ref_formulas = 0
        ccn_rows: list[tuple[int, str, str]] = []  # (row, ccn, sample_text)
        for r in range(1, max_r + 1):
            row_text_parts: list[str] = []
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if isinstance(v, str):
                    if v.startswith("="):
                        formula_count += 1
                        if "[" in v or "!" in v and ":" in v:
                            ext_ref_formulas += 1
                    row_text_parts.append(v)
                elif v is not None:
                    row_text_parts.append(str(v))
            row_text = " | ".join(row_text_parts[:8])
            m = CCN_RE.search(row_text)
            if m:
                ccn_rows.append((r, m.group(0), row_text[:140]))

        out.append(f"  formulas      : {formula_count}  (with ext refs: {ext_ref_formulas})")
        out.append(f"  ccn-like rows : {len(ccn_rows)}  (showing first 12)")
        for r, ccn, sample in ccn_rows[:12]:
            out.append(f"      row {r:>4}  CCN~{ccn}  :: {sample}")

    return "\n".join(out)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: inventory.py <xlsx> [<xlsx>...]", file=sys.stderr)
        sys.exit(2)
    for arg in sys.argv[1:]:
        print(inventory(Path(arg)))
        print("\n" + "#" * 72 + "\n")
