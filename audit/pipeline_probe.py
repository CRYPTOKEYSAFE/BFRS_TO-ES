"""Lock down the CCN+suffix tagging mechanism that links a billet to a facility CCN.

Hypothesis from inspecting the broken hidden CCN sheets:
    COUNTIFS(<TO!somecolumn>, "21710o", <TO!othercolumn>, B32)
    SUMIFS(<TE!somecolumn>, ..., "21710", ...)

The lookup string "21710o" must appear somewhere in either:
    - the in-workbook TO sheet of the CLB-4 SW BFR (which exists but may be partially populated),
    - the in-workbook TO sheet of 3D Med BN (which the schema map showed has a NOTE column),
    - or one of the company T/O files (Format A).

This script searches ALL TO/TE-style sheets across all unit data files for the
distinctive CCN+suffix tags and reports which column carries them.
"""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

CCN_SUFFIX_RE = re.compile(r"^\d{4,5}[a-z]{1,4}$")  # e.g. 21710o, 44112w, 21710rs


def scan(path: Path, sheet: str) -> None:
    print(f"\n## {path.name}  ::  {sheet}")
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as e:
        print(f"  load failed: {e}")
        return
    if sheet not in wb.sheetnames:
        print(f"  sheet not found")
        return
    ws = wb[sheet]
    col_hits: dict[int, Counter] = defaultdict(Counter)
    bare_ccn_hits: dict[int, Counter] = defaultdict(Counter)  # bare 5-digit CCNs
    total_rows = 0
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        total_rows = r
        for c, v in enumerate(row, start=1):
            if v is None:
                continue
            s = str(v).strip()
            if CCN_SUFFIX_RE.match(s):
                col_hits[c][s] += 1
            elif re.match(r"^\d{5}$", s):
                bare_ccn_hits[c][s] += 1
    print(f"  rows scanned: {total_rows}")
    if col_hits:
        print(f"  CCN+SUFFIX hits by column:")
        for c, counter in sorted(col_hits.items()):
            top = counter.most_common(8)
            print(f"    col {get_column_letter(c)} (idx {c}): {sum(counter.values())} hits, "
                  f"{len(counter)} distinct  e.g. {top}")
    else:
        print(f"  no CCN+suffix tags found")
    if bare_ccn_hits:
        print(f"  BARE 5-digit CCN hits by column (top 3 cols only):")
        sorted_cols = sorted(bare_ccn_hits.items(), key=lambda kv: -sum(kv[1].values()))[:3]
        for c, counter in sorted_cols:
            top = counter.most_common(6)
            print(f"    col {get_column_letter(c)} (idx {c}): {sum(counter.values())} hits, "
                  f"{len(counter)} distinct  e.g. {top}")


if __name__ == "__main__":
    targets = [
        # Format A: per-company T/O exports
        ("M29111_HQ_CO_CLB-4.xlsx", "TO"),
        ("M29112_CLC_A_CLB-4.xlsx", "TO"),
        ("M29113_CLC_B_CLB-4.xlsx", "TO"),
        ("M29114_GS_CO_CLB-4.xlsx", "TO"),
        ("M29111_HQ_CO_CLB-4.xlsx", "Primary Only"),
        ("M29112_CLC_A_CLB-4.xlsx", "Primary Only"),
        # Format B: BFR-attached TO/TE
        ("SW_M29030_CLB4_BFR_2026-NWPCW167400L021.xlsx", "TO"),
        ("SW_M29030_CLB4_BFR_2026-NWPCW167400L021.xlsx", "TE"),
        ("M67400-FO-M13020 3D MED BN-22NOV2024.xlsx", "TO"),
        ("M67400-FO-M13020 3D MED BN-22NOV2024.xlsx", "TE"),
        # Format C: Master MEF TO&E
        ("2031 Master TO&E v1.1 - 20250411.xlsx", "TO_2"),
        ("2031 Master TO&E v1.1 - 20250411.xlsx", "TE_3"),
    ]
    for path, sheet in targets:
        scan(Path(path), sheet)
