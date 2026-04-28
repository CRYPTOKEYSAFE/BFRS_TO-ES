"""Extract column headers from any sheet, picking the row that is actually the
header (not blank top rows or merged title rows).

Heuristic: scan rows 1..15 and pick the row with the most distinct non-empty
string values that don't look like data (no all-digits, no obvious unit
identifiers like M followed by 5 digits, etc.). Print all candidate rows so
a human can confirm.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook


M_UIC = re.compile(r"^M\d{5}$")


def looks_like_header(row: tuple) -> tuple[int, list[str]]:
    """Return (header_score, non_empty_labels)."""
    labels: list[str] = []
    for v in row:
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        labels.append(s)
    if not labels:
        return (0, labels)
    score = 0
    for s in labels:
        if s.isdigit():
            score -= 1
        elif M_UIC.match(s):
            score -= 2
        elif len(s) <= 40 and not s.startswith("="):
            score += 1
    return (score, labels)


def find_header(path: Path, sheet: str, max_scan: int = 15) -> None:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        print(f"  (sheet {sheet!r} not found)")
        return
    ws = wb[sheet]
    print(f"\n## {path.name}  ::  sheet={sheet!r}  (max_row~{ws.max_row}, max_col~{ws.max_column})")
    candidates: list[tuple[int, int, list[str]]] = []
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        score, labels = looks_like_header(row)
        if labels:
            candidates.append((r, score, labels))
    if not candidates:
        print("  no candidate rows")
        return
    best = max(candidates, key=lambda t: t[1])
    for r, score, labels in candidates:
        marker = "  <-- best" if (r, score, labels) == best else ""
        preview = " | ".join(labels[:8])
        if len(labels) > 8:
            preview += f" | ... ({len(labels)} total)"
        print(f"  row {r:>2} score={score:>3} : {preview}{marker}")
    # Print the best row's full label list
    r, score, labels = best
    print(f"\n  best header row {r} full column list:")
    for i, lab in enumerate(labels, start=1):
        print(f"    {i:>2}. {lab}")


if __name__ == "__main__":
    targets: list[tuple[str, str]] = [
        ("M29111_HQ_CO_CLB-4.xlsx", "TO"),
        ("M29112_CLC_A_CLB-4.xlsx", "TO"),
        ("M29113_CLC_B_CLB-4.xlsx", "TO"),
        ("M29114_GS_CO_CLB-4.xlsx", "TO"),
        ("M67400-FO-M13020 3D MED BN-22NOV2024.xlsx", "TO"),
        ("2031 Master TO&E v1.1 - 20250411.xlsx", "TO_2"),
        # And the equipment-side sheets
        ("M29111_HQ_CO_CLB-4.xlsx", "Primary Only"),
        ("M29112_CLC_A_CLB-4.xlsx", "Primary Only"),
        ("M29113_CLC_B_CLB-4.xlsx", "Primary Only"),
        ("M29114_GS_CO_CLB-4.xlsx", "Primary Only"),
        ("M67400-FO-M13020 3D MED BN-22NOV2024.xlsx", "TE"),
        ("2031 Master TO&E v1.1 - 20250411.xlsx", "TE_3"),
    ]
    for path_str, sheet in targets:
        find_header(Path(path_str), sheet)
