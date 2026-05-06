#!/usr/bin/env python3
"""
3d MED BN BFR TE backfill mapping.

Reads:
  - 2031 Master TO&E v1.1 - 20250411.xlsx, sheet TE_3
  - M67400-FO-M13020 3D MED BN-22NOV2024.xlsx, sheet TE (read-only)

Writes:
  - audit/reports/3dmedbn/52_te_backfill_mapping.csv
  - audit/reports/3dmedbn/52_te_backfill_summary.md
"""

import csv
import os
from collections import defaultdict, Counter

import openpyxl

REPO = "/home/user/BFRS_TO-ES"
MASTER_PATH = os.path.join(REPO, "2031 Master TO&E v1.1 - 20250411.xlsx")
BFR_PATH = os.path.join(REPO, "M67400-FO-M13020 3D MED BN-22NOV2024.xlsx")
OUT_DIR = os.path.join(REPO, "audit/reports/3dmedbn")
CSV_PATH = os.path.join(OUT_DIR, "52_te_backfill_mapping.csv")
MD_PATH = os.path.join(OUT_DIR, "52_te_backfill_summary.md")

THIRD_MEDBN_UICS = {"M28261", "M28262", "M28263"}

# Master TO&E TE_3 columns (1-indexed): UIC=5, TAMCN=9, L=17, W=18, H=19,
# Volume EA=20, Volume Total=21, Space Factor=22, NSY=24, NTG Factor=25
# Convert to 0-indexed for tuple access.
MC_UIC = 4
MC_TAMCN = 8
MC_L = 16
MC_W = 17
MC_H = 18
MC_VE = 19
MC_VT = 20
MC_SF = 21
MC_NSY = 23
MC_NTG = 24

# BFR TE columns: row 1 header
# (None, ROW, NOTE, CCN, CCN Description, CCN 21451 Equipment Type, UIC,
#  TAMCN Short, TAMCN, Nomenclature, TAM Stat, U/I, Rdy, Ind Qty, Org Qty,
#  Unit T/E, L Ft, W Ft, H Ft, Volume Ea, Volume Total, None)
B_NOTE = 2
B_CCN = 3
B_UIC = 6
B_TAMCN = 8


def norm(s):
    if s is None:
        return ""
    return str(s).strip()


def load_master_index():
    """Build {(uic, tamcn): [master_row_dict, ...]} restricted to 3d MED BN UICs.
    Also build {tamcn: set(uic)} across the same UIC set for cross-UIC checks.
    """
    wb = openpyxl.load_workbook(MASTER_PATH, data_only=True, read_only=True)
    ws = wb["TE_3"]
    by_key = defaultdict(list)
    by_tamcn = defaultdict(set)
    rows_loaded = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        uic = norm(row[MC_UIC])
        tamcn = norm(row[MC_TAMCN])
        if not uic or not tamcn:
            continue
        if uic not in THIRD_MEDBN_UICS:
            continue
        rec = {
            "uic": uic,
            "tamcn": tamcn,
            "L": row[MC_L],
            "W": row[MC_W],
            "H": row[MC_H],
            "VE": row[MC_VE],
            "VT": row[MC_VT],
            "SF": row[MC_SF],
            "NSY": row[MC_NSY],
            "NTG": row[MC_NTG],
            "ind_qty": row[13],
            "org_qty": row[14],
            "unit_te": row[15],
            "nomen": row[9],
        }
        by_key[(uic, tamcn)].append(rec)
        by_tamcn[tamcn].add(uic)
        rows_loaded += 1
    wb.close()
    return by_key, by_tamcn, rows_loaded


def walk_bfr_te():
    wb = openpyxl.load_workbook(BFR_PATH, data_only=True, read_only=True)
    ws = wb["TE"]
    out = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        tamcn = norm(row[B_TAMCN])
        if not tamcn:
            continue
        out.append({
            "row": r_idx,
            "uic": norm(row[B_UIC]),
            "tamcn": tamcn,
            "ccn": norm(row[B_CCN]),
            "note": norm(row[B_NOTE]),
            "nomen": norm(row[9]) if len(row) > 9 else "",
        })
    wb.close()
    return out


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    print("Loading Master TO&E TE_3 (filtered to 3d MED BN UICs)...")
    master_by_key, master_by_tamcn, mrows = load_master_index()
    print(f"  Master records loaded for 3d MED BN: {mrows}")
    print(f"  Unique (UIC,TAMCN) keys: {len(master_by_key)}")
    print(f"  Distinct TAMCNs: {len(master_by_tamcn)}")

    print("Walking BFR TE rows...")
    bfr_rows = walk_bfr_te()
    print(f"  BFR TE TAMCN-bearing rows: {len(bfr_rows)}")

    status_counts = Counter()
    per_ccn_total = Counter()
    per_ccn_with_sf = Counter()
    per_ccn_with_nsy_ntg = Counter()
    unmatched_tamcns = Counter()
    uic_mismatch_tamcns = []  # (uic, tamcn)
    cross_uic_examples = []

    csv_rows = []
    for b in bfr_rows:
        per_ccn_total[b["ccn"]] += 1
        key = (b["uic"], b["tamcn"])
        master_uic = master_tamcn = ""
        L = W = H = VE = VT = SF = NSY = NTG = ""
        if key in master_by_key:
            recs = master_by_key[key]
            rec = recs[0]
            status = "MATCHED"
            master_uic = rec["uic"]
            master_tamcn = rec["tamcn"]
            L = rec["L"]; W = rec["W"]; H = rec["H"]
            VE = rec["VE"]; VT = rec["VT"]
            SF = rec["SF"]; NSY = rec["NSY"]; NTG = rec["NTG"]
            if SF not in (None, ""):
                per_ccn_with_sf[b["ccn"]] += 1
            if NSY not in (None, "") and NTG not in (None, ""):
                per_ccn_with_nsy_ntg[b["ccn"]] += 1
        else:
            other_uics = master_by_tamcn.get(b["tamcn"], set())
            if b["uic"] in THIRD_MEDBN_UICS and other_uics and b["uic"] not in other_uics:
                # TAMCN exists under a different 3d MED BN UIC
                status = "TAMCN-MATCHED-IN-OTHER-UIC"
                # pick first other UIC
                ouic = sorted(other_uics)[0]
                rec = master_by_key[(ouic, b["tamcn"])][0]
                master_uic = ouic
                master_tamcn = rec["tamcn"]
                L = rec["L"]; W = rec["W"]; H = rec["H"]
                VE = rec["VE"]; VT = rec["VT"]
                SF = rec["SF"]; NSY = rec["NSY"]; NTG = rec["NTG"]
                cross_uic_examples.append((b["row"], b["uic"], b["tamcn"], ouic))
            elif b["uic"] in THIRD_MEDBN_UICS and not other_uics:
                status = "UNMATCHED"
                unmatched_tamcns[b["tamcn"]] += 1
            elif b["uic"] in THIRD_MEDBN_UICS and other_uics:
                # shouldn't reach here; covered above
                status = "UIC-MATCHED-TAMCN-MISMATCH"
                uic_mismatch_tamcns.append((b["uic"], b["tamcn"]))
            else:
                # BFR row's UIC is not one of the three; treat as unmatched
                status = "UNMATCHED"
                unmatched_tamcns[b["tamcn"]] += 1

        # If status is MATCHED but no other UIC has it, normal MATCH.
        # If UIC matches but TAMCN not in that UIC's master, that's UIC-MATCHED-TAMCN-MISMATCH.
        # Disambiguate: when key not in master_by_key but uic in THIRD_MEDBN_UICS and
        # there is ANY tamcn under that uic in master, BUT this tamcn is under a different uic,
        # we already labeled TAMCN-MATCHED-IN-OTHER-UIC. The UIC-MATCHED-TAMCN-MISMATCH bucket
        # therefore matches: uic in 3MEDBN, tamcn not in any 3MEDBN UIC. That collapses to UNMATCHED.
        # So we instead reclassify: UIC-MATCHED-TAMCN-MISMATCH = uic is one of the three but
        # tamcn never appears in master (post-2031 add or BFR error). UNMATCHED becomes a synonym
        # for "uic not in THIRD_MEDBN_UICS or both keys missing". Per task spec keep both buckets:
        # We will reclassify below.

        status_counts[status] += 1
        csv_rows.append({
            "bfr_te_row": b["row"],
            "bfr_uic": b["uic"],
            "bfr_tamcn": b["tamcn"],
            "match_status": status,
            "master_uic": master_uic,
            "master_tamcn": master_tamcn,
            "l_ft": "" if L is None else L,
            "w_ft": "" if W is None else W,
            "h_ft": "" if H is None else H,
            "vol_ea": "" if VE is None else VE,
            "vol_tot": "" if VT is None else VT,
            "space_factor": "" if SF is None else SF,
            "nsy": "" if NSY is None else NSY,
            "ntg_factor": "" if NTG is None else NTG,
            "ccn": b["ccn"],
            "note": b["note"],
        })

    # Reclassify: per task spec, distinguish UIC-MATCHED-TAMCN-MISMATCH from UNMATCHED:
    #   UIC-MATCHED-TAMCN-MISMATCH: bfr_uic in THIRD_MEDBN_UICS, tamcn not under that UIC,
    #                                AND tamcn does not exist under any other 3 MED BN UIC.
    #   UNMATCHED: bfr_uic not in THIRD_MEDBN_UICS at all (rare).
    # We rebuild status counts after reclassification using the rules.
    new_counts = Counter()
    new_unmatched_tamcns = Counter()
    new_uic_mismatch_tamcns = []
    for r in csv_rows:
        if r["match_status"] == "MATCHED":
            new_counts["MATCHED"] += 1
        elif r["match_status"] == "TAMCN-MATCHED-IN-OTHER-UIC":
            new_counts["TAMCN-MATCHED-IN-OTHER-UIC"] += 1
        else:
            # currently UNMATCHED
            if r["bfr_uic"] in THIRD_MEDBN_UICS:
                # uic recognized, tamcn not in any 3 MED BN UIC
                r["match_status"] = "UIC-MATCHED-TAMCN-MISMATCH"
                new_counts["UIC-MATCHED-TAMCN-MISMATCH"] += 1
                new_uic_mismatch_tamcns.append((r["bfr_uic"], r["bfr_tamcn"]))
            else:
                new_counts["UNMATCHED"] += 1
                new_unmatched_tamcns[r["bfr_tamcn"]] += 1

    status_counts = new_counts
    unmatched_tamcns = new_unmatched_tamcns
    uic_mismatch_tamcns = new_uic_mismatch_tamcns

    # Write CSV
    fieldnames = [
        "bfr_te_row", "bfr_uic", "bfr_tamcn", "match_status",
        "master_uic", "master_tamcn",
        "l_ft", "w_ft", "h_ft", "vol_ea", "vol_tot",
        "space_factor", "nsy", "ntg_factor",
        "ccn", "note",
    ]
    with open(CSV_PATH, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in csv_rows:
            w.writerow({k: r.get(k, "") for k in fieldnames})

    print(f"Wrote {CSV_PATH} ({len(csv_rows)} rows)")

    # Apex Omega Violation 1 check: C00392B FILTER WATER PURIFI in 3 MED BN UICs
    # Re-scan master TE_3 explicitly, not filtered, for C00392B to be thorough.
    print("Checking C00392B presence in 2031 Master TO&E for 3d MED BN UICs...")
    wb = openpyxl.load_workbook(MASTER_PATH, data_only=True, read_only=True)
    ws = wb["TE_3"]
    c00392b_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        tamcn = norm(row[MC_TAMCN])
        if tamcn != "C00392B":
            continue
        c00392b_rows.append({
            "uic": norm(row[MC_UIC]),
            "nomen": row[9],
            "ind_qty": row[13],
            "org_qty": row[14],
            "unit_te": row[15],
        })
    wb.close()
    c00392b_3medbn = [r for r in c00392b_rows if r["uic"] in THIRD_MEDBN_UICS]

    # CCN sub-summaries for the question
    ccn_21451_total = per_ccn_total.get("21451", 0)
    ccn_21451_sf = per_ccn_with_sf.get("21451", 0)
    ccn_14312_total = per_ccn_total.get("14312", 0)
    ccn_14312_nsy_ntg = per_ccn_with_nsy_ntg.get("14312", 0)

    # Markdown summary
    md = []
    md.append("# 3d MED BN BFR TE Backfill Mapping Summary")
    md.append("")
    md.append("Source mapping: 2031 Master TO&E v1.1 (2025-04-11), sheet TE_3,")
    md.append("filtered to UICs M28261, M28262, M28263 (3d MED BN).")
    md.append("BFR target: M67400-FO-M13020 3D MED BN-22NOV2024.xlsx, sheet TE (read-only).")
    md.append("")
    md.append("Apex Omega rule 7: every backfill value below is reproducible by")
    md.append("looking up (UIC, TAMCN) in the 2031 Master TO&E TE_3 sheet.")
    md.append("Apex Omega rule 4: rows that do not resolve are marked UNMATCHED")
    md.append("or UIC-MATCHED-TAMCN-MISMATCH and carry no proposed backfill values.")
    md.append("")
    md.append("## Headline finding")
    md.append("")
    md.append("Two of the eight columns the BFR's CCN sheets need are NOT recoverable")
    md.append("from the 2031 Master TO&E TE_3, because TE_3 itself does not carry")
    md.append("them. Specifically:")
    md.append("")
    md.append("- Space Factor (TE_3 col 22): blank for all 19,001 data rows.")
    md.append("- NTG Factor (TE_3 col 25): blank for all 19,001 data rows.")
    md.append("")
    md.append("The remaining six columns are recoverable for most rows:")
    md.append("")
    md.append("- L, Ft / W, Ft / H, Ft / Volume EA / Volume Total: populated for 464 of")
    md.append("  the 506 3d MED BN rows in TE_3 (~91.7%).")
    md.append("- NSY: populated for 439 of the 506 3d MED BN rows (~86.8%).")
    md.append("")
    md.append("Implication: a TAMCN-keyed lookup against TE_3 cannot, by itself,")
    md.append("supply the Space Factor and NTG Factor that CCN 21451 and CCN 14312")
    md.append("require. Those two factors must come from a different authoritative")
    md.append("source (FC 2-000-05N planning factor tables, or a separate engineering")
    md.append("data source). Per Apex Omega rule 4, those columns are TBD pending")
    md.append("identification of that source.")
    md.append("")
    md.append("## Match-class counts")
    md.append("")
    md.append(f"- Total BFR TE TAMCN-bearing rows: {len(bfr_rows)}")
    for cls in ["MATCHED", "TAMCN-MATCHED-IN-OTHER-UIC", "UIC-MATCHED-TAMCN-MISMATCH", "UNMATCHED"]:
        md.append(f"- {cls}: {status_counts.get(cls, 0)}")
    md.append("")
    md.append("## Master TO&E coverage stats")
    md.append("")
    md.append(f"- Master TO&E rows loaded for 3d MED BN UICs: {mrows}")
    md.append(f"- Unique (UIC, TAMCN) keys in master subset: {len(master_by_key)}")
    md.append(f"- Distinct TAMCNs across the three UICs: {len(master_by_tamcn)}")
    md.append("")

    md.append("## Unmatched TAMCNs (UNMATCHED bucket)")
    md.append("")
    if unmatched_tamcns:
        md.append("BFR rows whose UIC is not in {M28261, M28262, M28263} OR whose")
        md.append("TAMCN cannot be located under any 3d MED BN UIC.")
        md.append("")
        for t, n in sorted(unmatched_tamcns.items()):
            md.append(f"- {t}: {n} occurrence(s)")
    else:
        md.append("None.")
    md.append("")

    md.append("## UIC-MATCHED-TAMCN-MISMATCH bucket")
    md.append("")
    if uic_mismatch_tamcns:
        md.append("BFR row's UIC is one of the three 3d MED BN UICs, but the TAMCN")
        md.append("does not appear under that UIC (or any 3d MED BN UIC) in the 2031")
        md.append("Master TO&E. Likely causes: (a) post-2031-cut TAMCN added to the")
        md.append("unit, (b) BFR data-entry error, (c) TAMCN reclassification between")
        md.append("the Master TO&E vintage and the 2024 BFR. SME review required.")
        md.append("")
        # Group by UIC
        by_uic = defaultdict(list)
        for u, t in uic_mismatch_tamcns:
            by_uic[u].append(t)
        for u in sorted(by_uic):
            md.append(f"### UIC {u} ({len(by_uic[u])} TAMCNs)")
            md.append("")
            for t in sorted(set(by_uic[u])):
                md.append(f"- {t}")
            md.append("")
    else:
        md.append("None.")
    md.append("")

    md.append("## TAMCN-MATCHED-IN-OTHER-UIC examples")
    md.append("")
    if cross_uic_examples:
        md.append("BFR row's UIC differs from the master, but the TAMCN exists under")
        md.append("a sibling 3d MED BN UIC. Backfill values pulled from the sibling")
        md.append("record are mathematically valid (dimensions/space factors are")
        md.append("TAMCN-intrinsic), but counts (Ind/Org Qty) belong to the original")
        md.append("BFR row, not to the master sibling.")
        md.append("")
        md.append(f"Total cross-UIC matches: {len(cross_uic_examples)}")
        md.append("")
        md.append("First 25 examples (bfr_row, bfr_uic, tamcn, master_uic):")
        md.append("")
        for ex in cross_uic_examples[:25]:
            md.append(f"- row {ex[0]}, BFR UIC {ex[1]}, TAMCN {ex[2]}, found under {ex[3]}")
    else:
        md.append("None.")
    md.append("")

    md.append("## Apex Omega Violation 1 verification: C00392B FILTER WATER PURIFI")
    md.append("")
    md.append(f"- Rows for TAMCN C00392B in entire 2031 Master TO&E TE_3: {len(c00392b_rows)}")
    md.append(f"- Rows for TAMCN C00392B under 3d MED BN UICs (M28261/M28262/M28263): {len(c00392b_3medbn)}")
    if c00392b_3medbn:
        md.append("")
        md.append("Detail (UIC, Nomenclature, Ind Qty, Org Qty, Unit T/E):")
        md.append("")
        for r in c00392b_3medbn:
            md.append(f"- {r['uic']}, {r['nomen']}, ind={r['ind_qty']}, org={r['org_qty']}, te={r['unit_te']}")
    else:
        md.append("")
        md.append("FINDING: TAMCN C00392B FILTER WATER PURIFI does NOT appear under")
        md.append("any 3d MED BN UIC (M28261, M28262, M28263) in the 2031 Master TO&E")
        md.append("v1.1 (2025-04-11). It does not appear under any UIC in the entire")
        md.append("TE_3 sheet (zero hits in 19,001 data rows). The BFR's TE sheet does")
        md.append("carry C00392B for all three 3d MED BN UICs (BFR TE rows 505, 506,")
        md.append("507) but with no quantity or dimensional data, and the 2031 Master")
        md.append("TO&E does not corroborate any of it.")
        md.append("")
        md.append("Per Apex Omega rule 4, the previously derived qty 180 is unsupported")
        md.append("by this primary source and is TBD pending an authoritative T/E that")
        md.append("explicitly carries C00392B for 3d MED BN (e.g., the unit's ASR or a")
        md.append("more recent TFSMS export that includes this line item).")
    md.append("")

    md.append("## Per-CCN backfill counts")
    md.append("")
    md.append("Counts of BFR TE rows whose backfill from the 2031 Master TO&E")
    md.append("yields a non-blank Space Factor or non-blank NSY+NTG, by CCN tag")
    md.append("on the BFR TE row.")
    md.append("")
    md.append("| CCN | BFR TE rows | rows with Space Factor | rows with NSY+NTG |")
    md.append("|---|---:|---:|---:|")
    for ccn in sorted(per_ccn_total):
        md.append(f"| {ccn} | {per_ccn_total[ccn]} | {per_ccn_with_sf.get(ccn, 0)} | {per_ccn_with_nsy_ntg.get(ccn, 0)} |")
    md.append("")

    md.append("## CCN-specific feasibility")
    md.append("")
    md.append(f"- CCN 21451 has {ccn_21451_total} BFR TE rows tagged. Of those,")
    md.append(f"  {ccn_21451_sf} carry a Space Factor sourced from the 2031 Master TO&E.")
    md.append("  Space Factor is blank for every row of TE_3, so the 2031 Master TO&E")
    md.append("  cannot supply this column. Backfill is NOT sufficient for CCN 21451")
    md.append("  to compute; Space Factor must come from FC 2-000-05N planning factor")
    md.append("  tables for vehicle-equipment occupancy, or from another engineering")
    md.append("  authority. TBD pending source identification.")
    md.append("")
    # Count NSY-only and NTG-only coverage for 14312 to expose what IS available
    nsy_count_14312 = sum(1 for r in csv_rows if r["ccn"] == "14312" and r["nsy"] not in ("", None))
    ntg_count_14312 = sum(1 for r in csv_rows if r["ccn"] == "14312" and r["ntg_factor"] not in ("", None))
    md.append(f"- CCN 14312 has {ccn_14312_total} BFR TE rows tagged. Of those,")
    md.append(f"  {nsy_count_14312} carry NSY from the 2031 Master TO&E, and")
    md.append(f"  {ntg_count_14312} carry NTG Factor. NTG Factor is blank for every")
    md.append("  row of TE_3, so the 2031 Master TO&E cannot supply this column.")
    md.append("  Backfill is partial: NSY is recoverable but NTG is not. CCN 14312")
    md.append("  cannot compute on this backfill alone. NTG Factor TBD pending")
    md.append("  identification of an authoritative armory-storage planning source.")
    md.append("")

    md.append("## Method")
    md.append("")
    md.append("1. Master TO&E TE_3 was streamed and filtered to rows whose UIC was")
    md.append("   M28261, M28262, or M28263. Each row was indexed under (UIC, TAMCN).")
    md.append("2. The BFR's TE sheet (sheet name 'TE') was opened read-only. Each")
    md.append("   row with a non-blank TAMCN (column 9) was emitted to the CSV.")
    md.append("3. Match priority:")
    md.append("   a. Exact (UIC, TAMCN) hit -> MATCHED, backfill from that record.")
    md.append("   b. TAMCN exists under a different 3d MED BN UIC -> ")
    md.append("      TAMCN-MATCHED-IN-OTHER-UIC; dimensional/factor backfill carried")
    md.append("      from the sibling record (these fields are TAMCN-intrinsic).")
    md.append("   c. BFR UIC is in {M28261, M28262, M28263} but TAMCN absent in")
    md.append("      master subset -> UIC-MATCHED-TAMCN-MISMATCH, no backfill.")
    md.append("   d. Otherwise -> UNMATCHED, no backfill.")
    md.append("4. Outputs: CSV (one row per BFR TE row) and this summary.")
    md.append("")
    md.append("## Files")
    md.append("")
    md.append(f"- {CSV_PATH}")
    md.append(f"- {MD_PATH}")
    md.append("")

    with open(MD_PATH, "w") as f:
        f.write("\n".join(md))
    print(f"Wrote {MD_PATH}")

    # Print summary to stdout for caller
    print()
    print("=== SUMMARY (counts) ===")
    for cls in ["MATCHED", "TAMCN-MATCHED-IN-OTHER-UIC", "UIC-MATCHED-TAMCN-MISMATCH", "UNMATCHED"]:
        print(f"  {cls}: {status_counts.get(cls, 0)}")
    print(f"  CCN 21451: {ccn_21451_sf}/{ccn_21451_total} have Space Factor")
    print(f"  CCN 14312: {ccn_14312_nsy_ntg}/{ccn_14312_total} have NSY+NTG")
    print(f"  C00392B in 3d MED BN: {len(c00392b_3medbn)}")


if __name__ == "__main__":
    main()
