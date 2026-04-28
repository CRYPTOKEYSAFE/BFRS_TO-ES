"""Dump the contents of a single sheet from an xlsx workbook.

Reads each cell twice: once to capture the formula (data_only=False),
once to capture the cached computed value (data_only=True). Emits a
text report showing every non-empty cell with both values side-by-side
so we can spot:
  - cells whose formula references #REF! but whose cached value looks normal
  - cells whose cached value is empty/zero because IFERROR is masking a break
  - cells with hardcoded values that should be formulas (or vice versa)
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def dump(path: Path, sheet_name: str, max_row: int | None = None) -> str:
    out: list[str] = []
    out.append(f"FILE : {path.name}")
    out.append(f"SHEET: {sheet_name!r}")
    out.append("")

    wb_f = load_workbook(filename=str(path), data_only=False, read_only=False)
    wb_v = load_workbook(filename=str(path), data_only=True, read_only=False)
    if sheet_name not in wb_f.sheetnames:
        return f"sheet {sheet_name!r} not found in {path.name}\n"
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]

    rows = ws_f.max_row or 0
    cols = ws_f.max_column or 0
    if max_row is not None:
        rows = min(rows, max_row)
    out.append(f"DIMS : {rows} rows x {cols} cols  (state={ws_f.sheet_state})")
    out.append("")

    out.append(f"{'CELL':<8} {'FORMULA / CONTENT':<70} | CACHED VALUE")
    out.append("-" * 100)
    n_emit = 0
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            cf = ws_f.cell(row=r, column=c).value
            cv = ws_v.cell(row=r, column=c).value
            if cf is None and cv is None:
                continue
            ref = f"{get_column_letter(c)}{r}"
            f_str = "" if cf is None else str(cf)
            v_str = "" if cv is None else str(cv)
            if len(f_str) > 70:
                f_str = f_str[:67] + "..."
            if len(v_str) > 30:
                v_str = v_str[:27] + "..."
            out.append(f"{ref:<8} {f_str:<70} | {v_str}")
            n_emit += 1
    out.append("")
    out.append(f"TOTAL non-empty cells emitted: {n_emit}")
    return "\n".join(out)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("usage: sheet_dump.py <xlsx> <sheet_name> [max_row]", file=sys.stderr)
        sys.exit(2)
    path = Path(sys.argv[1])
    sheet = sys.argv[2]
    max_row = int(sys.argv[3]) if len(sys.argv) > 3 else None
    print(dump(path, sheet, max_row))
