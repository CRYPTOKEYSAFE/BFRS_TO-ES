"""
Surgical edit of MCBJ_BFR_Generator_FC2-000-05N.xlsx, authorized by user
on 2026-04-28:

  1. Drop the `DD1391` defined name and any cells that reference it.
     BFRs do not carry DD 1391s; the DD 1391 is a downstream MILCON
     project document, not a BFR field.
  2. Drop legacy "MCBJ" and "COMMARCORBASESJAPAN" terminology wherever
     it appears as a place / organization term. Current correct
     designation is MCIPAC (Marine Corps Installations Pacific) at
     MCB Camp Butler. CLAUDE.md "Terminology" section is the binding
     rule.

The workbook layout, fonts, fills, borders, merged cells, page setup,
and named-range API (other than `DD1391`) are preserved unchanged.
This is text/value-only surgery; cosmetic structure is untouched.

Run:  python3 audit/strip_dd1391_and_mcbj.py
Then: libreoffice --headless --calc --convert-to xlsx ... (recalc)
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.workbook.defined_name import DefinedName

REPO_ROOT = Path(__file__).resolve().parent.parent
WB_PATH = REPO_ROOT / "MCBJ_BFR_Generator_FC2-000-05N.xlsx"

NEW_B61 = (
    '=("This Basic Facility Requirement supports "&TENANT_UNIT&" at "'
    '&INSTALLATION&" under MCIPAC. Total personnel loading is "'
    '&TEXT(PN_TOTAL,"#,##0")&" PN ("&TEXT(PN_MIL,"#,##0")&" military). '
    'Requirements are computed in accordance with FC 2-000-05N Series 100. '
    'Existing assets per iNFADS have been compared against the required '
    'allowance to identify deficits; programmed cost reflects the Okinawa '
    'MILCON Area Cost Factor of "&TEXT(Okinawa_Navy_ACF,"0.00")'
    '&" per UFS 3-701-01 Table 4-1. ATFP (UFC 4-010-01), seismic design '
    '(Okinawa moderate seismic / typhoon Vult ≈ 170 mph), and '
    'host-nation siting requirements have been reviewed in the Okinawa_Adj '
    'checklist.")'
)


def main():
    if not WB_PATH.exists():
        print(f"FATAL: missing {WB_PATH}", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.load_workbook(WB_PATH)
    changes = []

    if "DD1391" in wb.defined_names:
        del wb.defined_names["DD1391"]
        changes.append("defined name DD1391 -> deleted")

    cov = wb["Cover"]
    for coord in ("C15", "D15"):
        if cov[coord].value is not None:
            old = cov[coord].value
            cov[coord].value = None
            changes.append(f"Cover!{coord}: cleared (was {old!r})")

    sm = wb["BFR_Summary"]
    for coord in ("B15", "C15"):
        if sm[coord].value is not None:
            old = sm[coord].value
            sm[coord].value = None
            changes.append(f"BFR_Summary!{coord}: cleared (was {old!r})")

    old_b61 = sm["B61"].value
    sm["B61"].value = NEW_B61
    changes.append(
        "BFR_Summary!B61: rewritten (dropped DD1391 reference and "
        "COMMARCORBASESJAPAN/MCBJ legacy terminology)"
    )

    cell_subs = [
        ("Cover", "B3",
         "FC 2-000-05N · Series 100 — Operational & Training "
         "Facilities · MCBJ / MCIPAC · Okinawa, Japan",
         "FC 2-000-05N · Series 100 — Operational & Training "
         "Facilities · MCIPAC · Okinawa, Japan"),
        ("Cover", "D9",
         "MCIPAC / MCBJ (COMMARCORBASESJAPAN) — 3d MLG / CLR-3",
         "MCIPAC — 3d MLG / CLR-3"),
        ("Okinawa_Adj", "B2",
         "Okinawa / MCBJ Adjustments",
         "Okinawa / MCIPAC Adjustments"),
        ("BFR_Summary", "B73",
         "MCBJ G-F (Facilities)",
         "MCIPAC G-F (Facilities)"),
    ]
    for sheet, coord, expected_old, new in cell_subs:
        ws = wb[sheet]
        actual = ws[coord].value
        if actual == expected_old:
            ws[coord].value = new
            changes.append(f"{sheet}!{coord}: {expected_old!r} -> {new!r}")
        else:
            changes.append(
                f"{sheet}!{coord}: SKIPPED (value did not match expected "
                f"old; actual={actual!r})"
            )

    wb.save(WB_PATH)
    print(f"Saved: {WB_PATH}")
    print()
    print("Changes:")
    for c in changes:
        print(f"  {c}")


if __name__ == "__main__":
    main()
