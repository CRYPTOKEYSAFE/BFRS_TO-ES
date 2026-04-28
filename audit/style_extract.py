"""Extract the cosmetic fingerprint of one or more sheets.

Captures everything a generator would need to reproduce the look:
  - workbook-level: theme, default font
  - sheet-level: tab color, page setup (orientation, paper size, margins,
    print area, fit-to-page), header/footer, default row/col dims, freeze panes
  - per-column: width, hidden flag
  - per-row: height, hidden flag
  - per-cell (non-empty only): font name/size/bold/italic/color, fill
    pattern + fg/bg color, border (each of 4 sides: style + color),
    alignment (horizontal/vertical/wrap/indent), number format,
    protection
  - merged cell ranges

Output is a single text report per sheet, deterministic and diff-friendly.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def color_str(c) -> str:
    if c is None:
        return "-"
    parts = []
    for attr in ("type", "value", "tint", "indexed", "theme", "rgb"):
        v = getattr(c, attr, None)
        if v is not None and v != 0:
            parts.append(f"{attr}={v}")
    return "{" + ",".join(parts) + "}" if parts else "-"


def font_str(f) -> str:
    if f is None:
        return "-"
    return (f"name={f.name!r} sz={f.size} bold={f.bold} italic={f.italic} "
            f"underline={f.underline} color={color_str(f.color)} "
            f"family={f.family} scheme={f.scheme}")


def fill_str(fill) -> str:
    if fill is None:
        return "-"
    pt = getattr(fill, "patternType", None) or getattr(fill, "fill_type", None)
    fg = getattr(fill, "fgColor", None) or getattr(fill, "start_color", None)
    bg = getattr(fill, "bgColor", None) or getattr(fill, "end_color", None)
    return f"pattern={pt} fg={color_str(fg)} bg={color_str(bg)}"


def side_str(s) -> str:
    if s is None:
        return "-"
    if not getattr(s, "style", None) and not getattr(s, "color", None):
        return "-"
    return f"style={s.style} color={color_str(getattr(s, 'color', None))}"


def border_str(b) -> str:
    if b is None:
        return "-"
    return (f"L={side_str(b.left)} R={side_str(b.right)} "
            f"T={side_str(b.top)} B={side_str(b.bottom)}")


def align_str(a) -> str:
    if a is None:
        return "-"
    return (f"h={a.horizontal} v={a.vertical} wrap={a.wrap_text} "
            f"indent={a.indent} rot={a.text_rotation} shrink={a.shrink_to_fit}")


def extract(path: Path, sheet: str) -> str:
    out: list[str] = []
    wb = load_workbook(filename=str(path), data_only=False, read_only=False)
    if sheet not in wb.sheetnames:
        return f"sheet {sheet!r} not found in {path.name}"
    ws = wb[sheet]
    out.append(f"FILE  : {path.name}")
    out.append(f"SHEET : {sheet!r}  state={ws.sheet_state}  "
               f"tab_color={color_str(ws.sheet_properties.tabColor)}")
    out.append(f"DIM   : {ws.dimensions}  "
               f"max_row={ws.max_row}  max_col={ws.max_column}")

    # Page setup
    ps = ws.page_setup
    pm = ws.page_margins
    out.append("")
    out.append("PAGE SETUP:")
    out.append(f"  orientation={ps.orientation} paper={ps.paperSize} "
               f"scale={ps.scale} fit_to_page={ws.sheet_properties.pageSetUpPr.fitToPage} "
               f"fit_w={ps.fitToWidth} fit_h={ps.fitToHeight}")
    out.append(f"  margins L={pm.left} R={pm.right} T={pm.top} B={pm.bottom} "
               f"hdr={pm.header} ftr={pm.footer}")
    out.append(f"  print_area={ws.print_area}")
    out.append(f"  print_title_rows={ws.print_title_rows} cols={ws.print_title_cols}")
    out.append(f"  freeze_panes={ws.freeze_panes}")

    hf = ws.HeaderFooter
    out.append(f"  hdr odd_left={hf.oddHeader.left.text!r} center={hf.oddHeader.center.text!r} "
               f"right={hf.oddHeader.right.text!r}")
    out.append(f"  ftr odd_left={hf.oddFooter.left.text!r} center={hf.oddFooter.center.text!r} "
               f"right={hf.oddFooter.right.text!r}")

    # Default dims
    out.append("")
    out.append(f"DEFAULTS: row_height={ws.sheet_format.defaultRowHeight}  "
               f"col_width={ws.sheet_format.defaultColWidth}  "
               f"baseColWidth={ws.sheet_format.baseColWidth}")

    # Column widths
    out.append("")
    out.append("COLUMNS (non-default widths):")
    for col_letter, dim in sorted(ws.column_dimensions.items()):
        if dim.width is not None or dim.hidden:
            out.append(f"  {col_letter}: width={dim.width} hidden={dim.hidden} "
                       f"outline={dim.outlineLevel}")

    # Row heights
    out.append("")
    out.append(f"ROWS (non-default heights, first 80):")
    rd_items = sorted(ws.row_dimensions.items())[:80]
    for r, dim in rd_items:
        if dim.height is not None or dim.hidden:
            out.append(f"  row {r:>4}: height={dim.height} hidden={dim.hidden}")

    # Merged cells
    out.append("")
    out.append(f"MERGED RANGES ({len(ws.merged_cells.ranges)}):")
    for mr in sorted(str(r) for r in ws.merged_cells.ranges):
        out.append(f"  {mr}")

    # Cell-level styles (non-empty only)
    out.append("")
    out.append("CELL STYLES (non-empty cells only):")
    seen_styles: dict[str, list[str]] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            key = "|".join([
                font_str(cell.font),
                fill_str(cell.fill),
                border_str(cell.border),
                align_str(cell.alignment),
                f"nfmt={cell.number_format}",
            ])
            seen_styles.setdefault(key, []).append(cell.coordinate)

    out.append(f"  distinct style fingerprints: {len(seen_styles)}")
    for i, (key, cells) in enumerate(
        sorted(seen_styles.items(), key=lambda kv: -len(kv[1]))
    ):
        out.append(f"")
        out.append(f"   style #{i+1}  used by {len(cells)} cells ")
        # Print key parts on separate lines for readability
        for part in key.split("|"):
            out.append(f"    {part}")
        sample = cells[:8]
        out.append(f"    e.g. {sample}{' ...' if len(cells) > 8 else ''}")

    return "\n".join(out)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("usage: style_extract.py <xlsx> <sheet> [<sheet>...]", file=sys.stderr)
        sys.exit(2)
    path = Path(sys.argv[1])
    for sheet in sys.argv[2:]:
        print(extract(path, sheet))
        print("\n" + "#" * 72 + "\n")
