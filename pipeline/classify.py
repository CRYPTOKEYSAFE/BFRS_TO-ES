"""
Layer 3 classifier (Apex Omega).

Loads audit/CLASSIFICATION_RULES.yaml and applies it to a billet
record from a Format A T/O&E. Emits a NOTE tag of the form
CCN+suffix or an unclassified disposition.

Run:
  python3 pipeline/classify.py --to <Format-A.xlsx> [--unit-context unit.json]

Or import as a library:
  from pipeline.classify import Classifier, BilletRecord
  c = Classifier()
  result = c.classify(BilletRecord(bic="...", billet_description="...", ...),
                       unit_context={"admin_ccn": "61072"})

The classifier is deterministic. Same inputs always produce the same
NOTE tag. Doctrine changes go through audit/CLASSIFICATION_RULES.yaml,
not the code.

Apex Omega:
  Rule 1: Facts only. The classifier does not invent BMOS-to-CCN
    mappings; it reads them from the YAML rule table, where each
    rule carries a citation field.
  Rule 4: TBD-not-guess. Rule rows where tag_template is "TBD" or
    contains an unresolved placeholder cause the classifier to emit
    "unclassified" with the rule id and citation, never a guessed
    NOTE tag.
  Rule 6: Three-bucket separation. Confidence flag (high / low / none)
    distinguishes general-pattern observed rules from BMOS-specific
    TBD rules from unclassified.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import yaml

REPO_ROOT = Path(__file__).resolve().parent.parent
RULES_PATH = REPO_ROOT / "audit" / "CLASSIFICATION_RULES.yaml"
UNIT_TYPE_DEFAULTS_PATH = REPO_ROOT / "audit" / "UNIT_TYPE_DEFAULTS.yaml"


PAY_GRADE = {}
for i, abbr in enumerate(
    ["PVT", "PFC", "LCPL", "CPL", "SGT", "SSGT",
     "GYSGT", "MSGT", "MGYSGT"], start=1):
    PAY_GRADE[abbr] = ("E", i)
PAY_GRADE["1STSGT"] = ("E", 8)
PAY_GRADE["SGTMAJ"] = ("E", 9)
for i, abbr in enumerate(
    ["WO", "CWO2", "CWO3", "CWO4", "CWO5"], start=1):
    PAY_GRADE[abbr] = ("W", i)
for i, abbr in enumerate(
    ["2NDLT", "1STLT", "CAPT", "MAJ", "LTCOL", "COL",
     "BGEN", "MAJGEN", "LTGEN", "GEN"], start=1):
    PAY_GRADE[abbr] = ("O", i)
for i, abbr in enumerate(
    ["ENS", "LTJG", "LT", "LCDR", "CDR"], start=1):
    PAY_GRADE.setdefault(abbr, ("O", i))
PAY_GRADE["CAPT(USN)"] = ("O", 6)
PAY_GRADE["RDML"] = ("O", 7)
PAY_GRADE["RADM"] = ("O", 8)
PAY_GRADE["VADM"] = ("O", 9)
PAY_GRADE["ADM"] = ("O", 10)
NAVY_RATE_NUM = re.compile(r"^[A-Z]{2,3}([1-9])$")
NAVY_RATE_C = re.compile(r"^[A-Z]{2,3}C$")
NAVY_RATE_CS = re.compile(r"^[A-Z]{2,3}CS$")
NAVY_RATE_CM = re.compile(r"^[A-Z]{2,3}CM$")


def parse_pay_grade(alpha_grade):
    """Return ('E'|'W'|'O', int) or None.

    Handles USMC abbreviations (LTCOL, GYSGT, etc.), Navy officer
    abbreviations (ENS, LT, LCDR, etc.), and Navy enlisted rate codes
    (HM1, RP3, MAC, etc.). Empty / unknown values return None.
    """
    if not alpha_grade:
        return None
    s = str(alpha_grade).strip().upper().replace(" ", "")
    if not s:
        return None
    if s in PAY_GRADE:
        return PAY_GRADE[s]
    m = NAVY_RATE_NUM.match(s)
    if m:
        n = int(m.group(1))
        if n in (1, 2, 3):
            return ("E", 7 - n)
    if NAVY_RATE_C.match(s):
        return ("E", 7)
    if NAVY_RATE_CS.match(s):
        return ("E", 8)
    if NAVY_RATE_CM.match(s):
        return ("E", 9)
    return None


def pay_grade_to_int(pg):
    """Map ('E', 5) to integer 5. ('W', 1) to integer 1. Branch ignored."""
    if pg is None:
        return None
    return pg[1]


SECTION_KEYWORDS = [
    ("admin_or_hq", [
        "COMMANDING OFFICER", "EXECUTIVE OFFICER", "COMMAND ELEMENT",
        "COMMAND SENIOR ENL", "SGTMAJ", "ADJUTANT", "S-1", "S1 ",
        "S-2", "S2 ", "S-3", "S3 ", "S-4", "S4 ", "S-5", "S5 ",
        "ADMIN", "TAC CHAPLAIN", "RP SPECIALIST", "STAFF",
        "PERSONNEL", "OPS", "OPERATIONS", "PLANS", "TRAINING",
        "BILLETING", "POSTAL", "MANPOWER", "INTELLIGENCE",
        "INTEL ", "FIRST SERGEANT", "SERGEANT MAJOR",
        "SENIOR ENLISTED", "UNIT LEADER"]),
    ("auto_org_shop", [
        "MOTOR T", "MOTOR TRANSPORT", "VEHICLE MAINT",
        "AUTO ORG", "AUTOMOTIVE", "DISPATCHER"]),
    ("comm_shop", [
        "COMMUNICATIONS", "COMM SECTION", "RADIO", "ANTENNA",
        "TELEPHONE", "WIRE", "SHF", "SATELLITE COMM"]),
    ("data_shop", [
        "DATA SYSTEMS", "S-6", "S6 ", "CYBER", "NETWORK",
        "INFORMATION SYSTEMS"]),
    ("medical", [
        "MEDIC", "CORPSMAN", "DENTAL", "DENT ", "DENT GP", "BAS ",
        "BATT AID", "HOSPITAL", "MEDICAL", "FLD MED", "FIELD MED",
        "NURSE", "NRS ", "TRAUMA", "IDC", "INDEPENDENT DUTY"]),
    ("eod", [
        "EOD", "EXPLOSIVE ORDNANCE DISP"]),
    ("ordnance", [
        "ORDNANCE", "AMMO", "AMMUNITION", "WEAPONS",
        "SMALL ARMS REPAIR", "ARMORER"]),
    ("supply", [
        "SUPPLY", "WAREHOUSE", "STORE", "LOGISTICS",
        "DISTRIBUTION", "MOBILITY", "MATERIEL"]),
    ("engineer", [
        "ENGINEER", "CONST", "CONSTRUCTION"]),
    ("utilities", [
        "ELECTRICIAN", "WATER SUPPORT", "UTILITIES SYSTEMS",
        "REFRIGERATION", "HVAC", "POWER GENERATION",
        "WASTEWATER", "POL ", "FUEL "]),
    ("mp", [
        "MILITARY POLICE", " MP ", "CORRECTIONS", "BRIG"]),
    ("field_maint", [
        "FIELD MAINT", "MAINTENANCE BAY", "RECOVERY",
        "MAINTENANCE TECHNICIAN", "MAINTENANCE CHIEF",
        "MAINTENANCE SNCOIC"]),
    ("food_service", [
        "FOOD SERVICE", "MESS", "GALLEY", "DINING", "COOK"]),
]


def infer_section(billet_description):
    """Return section name string or 'unknown'."""
    if not billet_description:
        return "unknown"
    s = " " + str(billet_description).upper() + " "
    for section, keywords in SECTION_KEYWORDS:
        for kw in keywords:
            if kw in s:
                return section
    return "unknown"


COMM_SUFFIX_KEYWORDS = [
    ("rs", ["RADIO SECTION", "RADIO OPER"]),
    ("cs", ["COMM SECTION", "COMMUNICATIONS SECTION"]),
    ("ds", ["DATA SECTION", "DATA SYSTEM"]),
    ("ws", ["WIRE SECTION"]),
    ("shf", ["SHF SECTION", "SATELLITE"]),
    ("ms", ["MAINT SECTION"]),
]


def infer_comm_section_suffix(billet_description):
    if not billet_description:
        return ""
    s = str(billet_description).upper()
    for suffix, keywords in COMM_SUFFIX_KEYWORDS:
        for kw in keywords:
            if kw in s:
                return suffix
    return ""


@dataclass
class BilletRecord:
    bic: str = ""
    billet_description: str = ""
    alpha_grade: str = ""
    bmos: str = ""
    pmos: str = ""
    mcc: str = ""
    uic: str = ""


@dataclass
class ClassifyResult:
    note_tag: Optional[str]
    confidence: str
    rule_id: str
    citation: str
    section_inferred: str
    pay_grade: Optional[str]
    unclassified_reason: Optional[str] = None


def _matches_rule(rule, billet, section, pg_int, pg_branch):
    when = rule.get("when", {})
    if "bmos_prefix" in when:
        prefix = when["bmos_prefix"]
        bmos = (billet.bmos or "").strip()
        if not bmos.startswith(prefix):
            return False
    if "alpha_grade_min" in when:
        target = when["alpha_grade_min"]
        target_pg = parse_pay_grade(target.replace("-", ""))
        if target_pg is None:
            target_pg = _parse_grade_dash(target)
        if target_pg is None or pg_int is None:
            return False
        if pg_branch != target_pg[0] or pg_int < target_pg[1]:
            return False
    if "alpha_grade_max" in when:
        target = when["alpha_grade_max"]
        target_pg = _parse_grade_dash(target)
        if target_pg is None or pg_int is None:
            return False
        if pg_branch != target_pg[0] or pg_int > target_pg[1]:
            return False
    if "section" in when and when["section"] != section:
        return False
    return True


def _parse_grade_dash(s):
    """Parse 'E-1', 'O-1', 'W-5' to ('E', 1) etc."""
    if not s:
        return None
    s = s.strip().upper()
    m = re.match(r"^([EOW])-?(\d+)$", s)
    if not m:
        return None
    return (m.group(1), int(m.group(2)))


def _resolve_template(template, context):
    """Substitute {placeholder} keys from context."""
    if template is None:
        return None
    out = template
    for key, val in context.items():
        out = out.replace("{" + key + "}", str(val) if val is not None else "")
    return out


class Classifier:

    def __init__(self, rules_path=RULES_PATH,
                 unit_type_defaults_path=UNIT_TYPE_DEFAULTS_PATH):
        if not rules_path.exists():
            raise FileNotFoundError(rules_path)
        with rules_path.open() as f:
            self.rules_doc = yaml.safe_load(f)
        self.defaults = self.rules_doc.get("defaults", {}) or {}
        self.rules = self.rules_doc.get("rules", []) or []
        self.unclassified = (
            self.rules_doc.get("unclassified_disposition", {}) or {}
        )
        # Layer 5 / Track 6: per-unit-type defaults. Loaded if present;
        # absent file is non-fatal (legacy callers without unit_type
        # still work, falling back to rule-table defaults).
        self.unit_type_defaults = {}
        self.unit_type_default_fallback = {}
        if unit_type_defaults_path.exists():
            with unit_type_defaults_path.open() as f:
                doc = yaml.safe_load(f) or {}
            self.unit_type_defaults = doc.get("unit_types", {}) or {}
            self.unit_type_default_fallback = doc.get("default", {}) or {}

    def _resolve_unit_type_defaults(self, unit_type):
        """Return the dict of defaults for the given unit_type, or the
        fallback dict if unit_type is None or not in the table. Apex
        Omega rule 4: when unit_type is unknown the fallback's
        admin_ccn is "TBD"; the classifier surfaces affected billets
        as orphans rather than substituting a guess."""
        if not unit_type:
            return dict(self.unit_type_default_fallback)
        row = self.unit_type_defaults.get(unit_type)
        if row is None:
            return dict(self.unit_type_default_fallback)
        return dict(row)

    def classify(self, billet, unit_context=None):
        # Precedence (lowest to highest):
        #   self.defaults                     (rule-table defaults)
        #   per-unit-type row from UNIT_TYPE_DEFAULTS.yaml
        #   explicit unit_context overrides   (highest)
        ctx = dict(self.defaults)
        unit_type = (unit_context or {}).get("unit_type")
        ut_row = self._resolve_unit_type_defaults(unit_type)
        # Only carry through fields the classifier knows how to use;
        # admin_ccn drops through to the rule template substitution.
        if "admin_ccn" in ut_row and ut_row["admin_ccn"] != "TBD":
            ctx["admin_ccn"] = ut_row["admin_ccn"]
        elif "admin_ccn" in ut_row and ut_row["admin_ccn"] == "TBD":
            # Apex Omega rule 4: do not silently inherit the
            # rule-table default 61072 when the unit_type's
            # admin_ccn is explicitly TBD. Mark it so the rule
            # template substitution leaves a placeholder, which
            # surfaces the billet as unclassified.
            ctx["admin_ccn"] = "{admin_ccn_TBD}"
        if unit_context:
            ctx.update(unit_context)
        # unit_type itself is a context-only key, not a template var;
        # remove it before template substitution to avoid noise.
        ctx.pop("unit_type", None)
        section = infer_section(billet.billet_description)
        if "section" not in ctx:
            ctx["section"] = section
        ctx.setdefault("comm_section_suffix",
                       infer_comm_section_suffix(billet.billet_description))
        pg = parse_pay_grade(billet.alpha_grade)
        pg_branch = pg[0] if pg else None
        pg_int = pay_grade_to_int(pg) if pg else None

        for rule in self.rules:
            if _matches_rule(rule, billet, section, pg_int, pg_branch):
                template = rule.get("tag_template")
                if not template or template == "TBD" or "TBD" in str(template):
                    return ClassifyResult(
                        note_tag=None,
                        confidence=rule.get("confidence", "none"),
                        rule_id=rule.get("id", ""),
                        citation=rule.get("citation", ""),
                        section_inferred=section,
                        pay_grade=f"{pg[0]}-{pg[1]}" if pg else None,
                        unclassified_reason=(
                            f"rule {rule.get('id')} matched but "
                            f"tag_template is TBD"
                        ),
                    )
                resolved = _resolve_template(template, ctx)
                if "{" in resolved or not resolved:
                    return ClassifyResult(
                        note_tag=None,
                        confidence=rule.get("confidence", "none"),
                        rule_id=rule.get("id", ""),
                        citation=rule.get("citation", ""),
                        section_inferred=section,
                        pay_grade=f"{pg[0]}-{pg[1]}" if pg else None,
                        unclassified_reason=(
                            f"rule {rule.get('id')} matched but template "
                            f"has unresolved placeholder: {template}"
                        ),
                    )
                return ClassifyResult(
                    note_tag=resolved,
                    confidence=rule.get("confidence", "high"),
                    rule_id=rule.get("id", ""),
                    citation=rule.get("citation", ""),
                    section_inferred=section,
                    pay_grade=f"{pg[0]}-{pg[1]}" if pg else None,
                )

        return ClassifyResult(
            note_tag=None,
            confidence="none",
            rule_id="",
            citation="",
            section_inferred=section,
            pay_grade=f"{pg[0]}-{pg[1]}" if pg else None,
            unclassified_reason="no rule matched",
        )


def read_format_a_to(workbook_path):
    """Read a Format-A T/O&E TO sheet, yield BilletRecord per data row.

    Filters TFSMS section-header rows and placeholder rows that appear
    in some Format-A exports (e.g., rows where the company structure
    is denoted by a description like "S-1", "COMMAND ELEMENT", or
    "EOD SECTION" with no Alpha Grade and no BMOS, or vacant-billet
    placeholder rows where every field except a coded BMOS suffix is
    empty). These are not real billets and would otherwise land as
    Apex Omega rule 4 orphans noisily."""
    import openpyxl
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if "TO" not in wb.sheetnames:
        return
    ws = wb["TO"]
    for r in range(8, ws.max_row + 1):
        bic = ws.cell(row=r, column=2).value
        billet_desc = ws.cell(row=r, column=4).value
        alpha = ws.cell(row=r, column=9).value
        bmos = ws.cell(row=r, column=11).value
        pmos = ws.cell(row=r, column=12).value
        if not (bic or alpha or bmos):
            continue
        # Filter TFSMS organizational divider rows: a real billet
        # always carries either a pay grade (filled or "VACANT") or
        # an MOS code. Rows where both are empty are section /
        # platoon / team / company-HQ headers in the export
        # (e.g., "S-1", "EOD SECTION", "MAINTENANCE PLATOON",
        # "DISTRIBUTION PLATOON", "COMPANY HEADQUARTERS"), not
        # billets, and would otherwise land as Apex Omega rule 4
        # orphans noisily.
        desc_str = (str(billet_desc).strip().upper() if billet_desc else "")
        if (not alpha) and (not bmos):
            continue
        # Filter placeholder rows where everything except a coded
        # BMOS suffix (e.g. "0113N", "8057D") is empty.
        if (not alpha) and (not desc_str) and bmos:
            bmos_str = str(bmos).strip()
            if re.match(r"^\d{3,4}[A-Z]$", bmos_str):
                continue
        yield BilletRecord(
            bic=str(bic) if bic else "",
            billet_description=str(billet_desc) if billet_desc else "",
            alpha_grade=str(alpha) if alpha else "",
            bmos=str(bmos) if bmos else "",
            pmos=str(pmos) if pmos else "",
            uic=str(workbook_path.stem.split("_")[0])
                if workbook_path.stem else "",
        )


def main():
    ap = argparse.ArgumentParser(description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--to", type=Path, required=True,
        help="Format-A T/O&E xlsx path")
    ap.add_argument("--unit-context", type=Path, default=None,
        help="JSON file with unit-context overrides (admin_ccn etc.)")
    ap.add_argument("--out", type=Path, default=None,
        help="CSV output path (defaults to stdout)")
    args = ap.parse_args()
    unit_ctx = {}
    if args.unit_context:
        unit_ctx = json.loads(args.unit_context.read_text())
    classifier = Classifier()
    rows = []
    for billet in read_format_a_to(args.to):
        result = classifier.classify(billet, unit_ctx)
        rows.append({
            "BIC": billet.bic,
            "BilletDescription": billet.billet_description,
            "AlphaGrade": billet.alpha_grade,
            "BMOS": billet.bmos,
            "Section": result.section_inferred,
            "PayGrade": result.pay_grade or "",
            "NOTE": result.note_tag or "",
            "Confidence": result.confidence,
            "RuleID": result.rule_id,
            "Citation": result.citation,
            "Unclassified": result.unclassified_reason or "",
        })
    cols = list(rows[0].keys()) if rows else []
    out_stream = open(args.out, "w", newline="") if args.out else sys.stdout
    w = csv.DictWriter(out_stream, fieldnames=cols, quoting=csv.QUOTE_ALL)
    w.writeheader()
    for row in rows:
        w.writerow(row)
    if args.out:
        out_stream.close()
    classified = sum(1 for r in rows if r["NOTE"])
    sys.stderr.write(
        f"\n{len(rows)} billets read; {classified} classified, "
        f"{len(rows) - classified} unclassified.\n")


if __name__ == "__main__":
    main()
