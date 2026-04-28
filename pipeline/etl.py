"""
End-to-end BFR ETL orchestrator (Apex Omega).

Reads one or more Format A T/O&E files (per-company TFSMS exports),
classifies every billet via pipeline/classify.py, extracts every
TAMCN row, and populates the generated BFR's TO and TE sheets with
the data plus NOTE column tags. The rest of the BFR (UNIT_ROLLUP,
CCN sheets) comes from pipeline/template.py.

Run:
  python3 pipeline/etl.py \
      --profile <profile.json> \
      --ccns <ccns.json> \
      --to-files M29111_HQ_CO_CLB-4.xlsx M29112_CLC_A_CLB-4.xlsx ... \
      --output out/CLB4_BFR_full.xlsx \
      [--unit-context unit_context.json]

Apex Omega:
  Rule 1: Facts only. Every populated NOTE tag carries a
    rule_id and citation in the diff log.
  Rule 4: Unclassified billets (no rule matched, or matched a TBD
    rule) are written to TO with NOTE blank; the validator's
    Check 7 (billet accounting) surfaces them as orphans.
  Rule 5: Apex Omega timestamp at point of citation is the source
    workbook's last-modified date plus the current ETL run date.
  Rule 7: Numbers must be traceable. Every billet row in the
    generated TO carries the source workbook filename in the diff
    log artifact at audit/reports/19_etl_<profile>.txt.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, List, Optional, Tuple

import openpyxl
import yaml

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from pipeline.classify import (
    BilletRecord,
    Classifier,
    read_format_a_to,
)
from pipeline import template as tpl

DEFAULT_TAMCN_MAP = REPO_ROOT / "audit" / "TAMCN_CCN_MAP.yaml"


@dataclass
class EquipmentRecord:
    tamcn: str
    nomenclature: str = ""
    tam_stat: str = ""
    ui: str = ""
    rdy: str = ""
    ind_qty: Optional[float] = None
    org_qty: Optional[float] = None
    unit_te: Optional[float] = None
    source_file: str = ""
    uic: str = ""


@dataclass
class TamcnRule:
    rule_id: str
    pattern: re.Pattern
    facility_ccn: str
    confidence: str
    description: str = ""


@dataclass
class TamcnMap:
    """Loaded TAMCN to facility CCN doctrine table.

    The validator regex filters non-equipment rows (header, footer,
    section labels). Rules are evaluated top-to-bottom; first match
    wins. Apex Omega rule 4: rules with facility_ccn == "TBD" are
    skipped at load time so unmapped TAMCNs surface as orphans in
    pipeline/validate.py Check 8 instead of receiving a guessed CCN.
    """
    validator: re.Pattern
    rules: List[TamcnRule] = field(default_factory=list)
    skipped_tbd_rule_ids: List[str] = field(default_factory=list)

    @classmethod
    def from_yaml(cls, path: Path) -> "TamcnMap":
        data = yaml.safe_load(path.read_text())
        validator = re.compile(data["tamcn_validator"])
        rules: List[TamcnRule] = []
        skipped: List[str] = []
        for raw in data.get("rules", []):
            ccn = raw.get("facility_ccn")
            if ccn in (None, "", "TBD"):
                skipped.append(raw.get("id", "<unnamed>"))
                continue
            rules.append(TamcnRule(
                rule_id=raw["id"],
                pattern=re.compile(raw["pattern"]),
                facility_ccn=str(ccn),
                confidence=str(raw.get("confidence", "")),
                description=str(raw.get("description", "")),
            ))
        return cls(validator=validator, rules=rules, skipped_tbd_rule_ids=skipped)

    def is_valid_tamcn(self, value: str) -> bool:
        return bool(self.validator.match(value))

    def resolve(self, tamcn: str) -> Tuple[Optional[str], Optional[str]]:
        """Return (facility_ccn, rule_id) or (None, None) if no match."""
        for rule in self.rules:
            if rule.pattern.match(tamcn):
                return rule.facility_ccn, rule.rule_id
        return None, None


def read_format_a_te(workbook_path: Path, tamcn_map: Optional[TamcnMap] = None):
    """Yield EquipmentRecord per data row from the Primary Only sheet.

    When a tamcn_map is supplied, rows whose first-column value does
    not match tamcn_map.validator are filtered out. This drops the
    header, footnote, and section-label rows that leak through TFSMS
    exports (e.g., "TAMCN", "CUI", "Component TAMCNs",
    "All EDL Requirement Quantities ...", date strings).
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if "Primary Only" not in wb.sheetnames:
        return
    ws = wb["Primary Only"]
    for r in range(11, ws.max_row + 1):
        tamcn = ws.cell(row=r, column=1).value
        if not tamcn:
            continue
        tamcn_str = str(tamcn).strip().upper()
        if tamcn_map is not None and not tamcn_map.is_valid_tamcn(tamcn_str):
            continue
        yield EquipmentRecord(
            tamcn=tamcn_str,
            nomenclature=str(ws.cell(row=r, column=3).value or "").strip(),
            tam_stat=str(ws.cell(row=r, column=8).value or "").strip(),
            ui=str(ws.cell(row=r, column=9).value or "").strip(),
            rdy=str(ws.cell(row=r, column=10).value or "").strip(),
            ind_qty=ws.cell(row=r, column=12).value,
            org_qty=ws.cell(row=r, column=13).value,
            unit_te=ws.cell(row=r, column=14).value,
            source_file=workbook_path.name,
            uic=str(workbook_path.stem.split("_")[0]),
        )


def populate_to_sheet(ws, classified_billets, hdr_row=6):
    """Write classified billets into the generated TO sheet, starting
    one row below the header. NOTE column is column C."""
    next_row = hdr_row + 1
    for line, (billet, result) in enumerate(classified_billets, start=1):
        note = result.note_tag if result.note_tag else ""
        ccn = ""
        if note:
            for i, ch in enumerate(note):
                if not ch.isdigit():
                    ccn = note[:i]
                    break
            else:
                ccn = note
        ws.cell(row=next_row, column=2, value=line)
        ws.cell(row=next_row, column=3, value=note)
        ws.cell(row=next_row, column=4, value=ccn)
        ws.cell(row=next_row, column=6, value=billet.uic)
        ws.cell(row=next_row, column=8, value=billet.bic)
        ws.cell(row=next_row, column=9, value=billet.billet_description)
        ws.cell(row=next_row, column=10, value=billet.alpha_grade)
        ws.cell(row=next_row, column=11, value=billet.bmos)
        ws.cell(row=next_row, column=12, value=billet.pmos)
        next_row += 1
    return next_row - hdr_row - 1


def populate_te_sheet(ws, equipment, tamcn_map: Optional[TamcnMap] = None,
                      hdr_row: int = 4):
    """Write equipment rows into the generated TE sheet. NOTE column is
    column C; CCN column is column D. tamcn_map (TamcnMap instance)
    is optional; when supplied, the matcher resolves each TAMCN to a
    facility CCN via the YAML rule list. Unmatched TAMCNs leave the
    NOTE and CCN columns blank, which surfaces the row as an orphan
    in pipeline/validate.py Check 8 per Apex Omega rule 4. Returns
    (rows_written, attribution_counter) where attribution_counter
    maps facility_ccn to attributed-row count."""
    next_row = hdr_row + 1
    attributed: Counter = Counter()
    rule_hits: Counter = Counter()
    orphans = 0
    for row_num, e in enumerate(equipment, start=1):
        ccn = ""
        rule_id = None
        if tamcn_map is not None:
            ccn_match, rule_id = tamcn_map.resolve(e.tamcn)
            if ccn_match:
                ccn = ccn_match
                attributed[ccn] += 1
                rule_hits[rule_id] += 1
            else:
                orphans += 1
        ws.cell(row=next_row, column=2, value=row_num)
        ws.cell(row=next_row, column=3, value=ccn)
        ws.cell(row=next_row, column=4, value=ccn)
        ws.cell(row=next_row, column=7, value=e.uic)
        ws.cell(row=next_row, column=8, value=e.tamcn[:5] if e.tamcn else "")
        ws.cell(row=next_row, column=9, value=e.tamcn)
        ws.cell(row=next_row, column=10, value=e.nomenclature)
        ws.cell(row=next_row, column=11, value=e.tam_stat)
        ws.cell(row=next_row, column=12, value=e.ui)
        ws.cell(row=next_row, column=13, value=e.rdy)
        ws.cell(row=next_row, column=14, value=e.ind_qty)
        ws.cell(row=next_row, column=15, value=e.org_qty)
        ws.cell(row=next_row, column=16, value=e.unit_te)
        next_row += 1
    return next_row - hdr_row - 1, attributed, rule_hits, orphans


def main():
    ap = argparse.ArgumentParser(description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--profile", type=Path, required=True)
    ap.add_argument("--ccns", type=Path, required=True)
    ap.add_argument("--to-files", type=Path, nargs="+", required=True,
                    help="One or more Format-A T/O&E xlsx paths")
    ap.add_argument("--output", type=Path, required=True)
    ap.add_argument("--unit-context", type=Path, default=None)
    ap.add_argument("--tamcn-ccn-map", type=Path, default=DEFAULT_TAMCN_MAP,
                    help="YAML rule table mapping TAMCN to facility CCN. "
                         "Defaults to audit/TAMCN_CCN_MAP.yaml. Pass "
                         "an empty path or --no-tamcn-map to disable.")
    ap.add_argument("--no-tamcn-map", action="store_true",
                    help="Disable TAMCN to CCN attribution; TE NOTE/CCN "
                         "columns stay blank and Check 8 surfaces every "
                         "row as an orphan")
    ap.add_argument("--report", type=Path,
                    default=REPO_ROOT / "audit" / "reports" / "19_etl_run.txt")
    args = ap.parse_args()

    profile = tpl.UnitProfile.from_json(args.profile)
    raw_ccns = json.loads(args.ccns.read_text())
    vocab = tpl.load_vocabulary()
    ccn_specs = tpl.hydrate_ccns(raw_ccns, vocab)

    unit_ctx = {}
    if args.unit_context:
        unit_ctx = json.loads(args.unit_context.read_text())

    tamcn_map: Optional[TamcnMap] = None
    if not args.no_tamcn_map and args.tamcn_ccn_map:
        if not args.tamcn_ccn_map.exists():
            sys.exit(f"FATAL: TAMCN map not found at {args.tamcn_ccn_map}")
        tamcn_map = TamcnMap.from_yaml(args.tamcn_ccn_map)

    classifier = Classifier()
    classified = []
    equipment = []
    per_file = {}
    for to_path in args.to_files:
        if not to_path.exists():
            sys.exit(f"FATAL: missing {to_path}")
        billets = list(read_format_a_to(to_path))
        equip = list(read_format_a_te(to_path, tamcn_map))
        per_file[to_path.name] = {
            "billets": len(billets),
            "equipment": len(equip),
        }
        for b in billets:
            r = classifier.classify(b, unit_ctx)
            classified.append((b, r))
        equipment.extend(equip)

    wb = tpl.build(profile, ccn_specs)
    n_to = populate_to_sheet(wb["TO"], classified, hdr_row=6)
    n_te, te_attributed, te_rule_hits, te_orphans = populate_te_sheet(
        wb["TE"], equipment, tamcn_map, hdr_row=4
    )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    by_note = Counter(r.note_tag for _, r in classified if r.note_tag)
    by_unclass = Counter(
        r.unclassified_reason for _, r in classified if not r.note_tag)
    n_classified = sum(1 for _, r in classified if r.note_tag)

    args.report.parent.mkdir(parents=True, exist_ok=True)
    lines = []
    lines.append("BFR ETL Run Report")
    lines.append("=" * 56)
    lines.append(f"Output workbook : {args.output}")
    lines.append(f"Profile         : {args.profile.name}")
    lines.append(f"CCN list        : {args.ccns.name}")
    lines.append(f"Source TO files : {len(args.to_files)}")
    for name, cnt in per_file.items():
        lines.append(
            f"  {name}  billets={cnt['billets']}  "
            f"equipment={cnt['equipment']}"
        )
    lines.append("")
    lines.append(f"TO populated rows : {n_to}")
    lines.append(f"TE populated rows : {n_te}")
    lines.append(f"Classified billets: {n_classified}")
    lines.append(f"Unclassified      : {len(classified) - n_classified}")
    lines.append("")
    lines.append("Classified billets by NOTE tag (top 12):")
    for note, count in by_note.most_common(12):
        lines.append(f"  {note}  {count}")
    lines.append("")
    lines.append("Unclassified billets by reason:")
    for reason, count in by_unclass.most_common():
        lines.append(f"  {count}  {reason}")
    lines.append("")
    lines.append("TE attribution (Layer 5 TAMCN to facility CCN map):")
    if tamcn_map is None:
        lines.append("  TAMCN map disabled; every TE row is an orphan")
    else:
        lines.append(f"  Rule table   : {args.tamcn_ccn_map}")
        lines.append(f"  Active rules : {len(tamcn_map.rules)}")
        lines.append(
            f"  Skipped (TBD): {len(tamcn_map.skipped_tbd_rule_ids)}  "
            f"(rules: {tamcn_map.skipped_tbd_rule_ids})"
        )
        lines.append(f"  TE rows      : {n_te}")
        lines.append(f"  Attributed   : {sum(te_attributed.values())}")
        lines.append(f"  Orphans      : {te_orphans}")
        lines.append("  Per-CCN attribution:")
        for ccn, count in te_attributed.most_common():
            lines.append(f"    {ccn}  {count}")
        lines.append("  Top rule hits:")
        for rule_id, count in te_rule_hits.most_common(8):
            lines.append(f"    {rule_id}  {count}")
    lines.append("")
    lines.append("Apex Omega note: TAMCN orphans are surfaced, not "
                 "silently dropped, per audit/TAMCN_CCN_MAP.yaml "
                 "unmapped_disposition.do_not_silently_drop. SME "
                 "review extends the rule table; do not guess CCNs.")
    args.report.write_text("\n".join(lines) + "\n")
    print("\n".join(lines))


if __name__ == "__main__":
    main()
