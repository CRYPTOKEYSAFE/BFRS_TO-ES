"""
Layer 4 BFR template generator (Apex Omega).

Produces a Format-B BFR workbook for any unit. Cosmetic typography
follows audit/STYLE_GUIDE.md (extracted from one CLB-4 SW example
unit; the typography is the NAVFAC convention used across DoD BFRs)
plus the Apex Omega four-color cell-role palette (FFF8DC input,
EAF3F4 calc, DCE7C8 output, F8E2D6 warning). Output is recalc-clean
(full-column lookups, no IFERROR masking) and validator-clean per
pipeline/validate.py.

Per-pattern row layouts in this module are observed shapes from one
example CLB-4 unit, not canonical Layer 5 patterns.
TBD, pending: ratify pattern definitions against FC 2-000-05N
planning factor tables (Series 100/200/etc.) plus unit-type doctrine
for the unit being generated. CLB-4 is an example, not a gold
standard. Different unit types (CLB, MAG, MWHS, MEU, depot, training
command, etc.) can have different shapes, factors, and CCN selections.

Run:
  python3 pipeline/template.py \\
      --profile <profile.json> \\
      --ccns <ccns.json> \\
      --output <path.xlsx>

Profile JSON schema (all fields required, strings):
  {
    "uic": "M29030",
    "tenant_unit": "CLB 4",
    "installation": "MCB CAMP BUTLER",
    "installation_uic": "M67400",
    "region": "Okinawa, Japan",
    "planning_area": "SW",
    "building_no": "TBD",
    "planner": "TBD",
    "fy": "2026",
    "project_date": "2026-04-28"
  }

CCNs JSON schema (list of objects):
  [
    {
      "ccn": "14345",
      "loading": 310,
      "factor": 1.0,
      "ntg": 1.0,
      "loading_label": "Total Personnel (PN_MIL)",
      "notes": ""
    },
    ...
  ]
  facility_name and uom are looked up from audit/CCN_VOCABULARY.json.

Output workbook layout:
  Sheet 'UNIT_ROLLUP'  : title banner + per-CCN total table.
  Sheet 'TO'           : Format-B header row 6, empty data rows.
  Sheet 'TE'           : Format-B header row 4, empty data rows.
  Sheet '<CCN>'        : one per CCN in --ccns; banner rows 1-7,
                         data rows 8-13, subtotal row 15-16,
                         TOTAL REQUIREMENT cell at H37.

Apex Omega notes:
  Source of CCN dictionary: audit/CCN_VOCABULARY.json (1,059 CCNs from
  FC 2-000-05N Appendix A, 27-JUN-2019 generation).
  This script is read-only against source data and writes only to the
  --output path.
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass, field, fields
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins, PrintOptions

REPO_ROOT = Path(__file__).resolve().parent.parent
VOCAB_PATH = REPO_ROOT / "audit" / "CCN_VOCABULARY.json"
PLANNING_FACTORS_PATH = REPO_ROOT / "audit" / "PLANNING_FACTORS.json"

# Cached at first call so repeat lookups across many CCNs in one
# build don't reread the file.
_PF_CACHE: dict | None = None


def _load_planning_factors() -> dict:
    """Return the parsed PLANNING_FACTORS.json keyed by CCN. Returns
    empty dict if the file is missing (caller falls back to a TBD
    citation rather than crashing)."""
    global _PF_CACHE
    if _PF_CACHE is not None:
        return _PF_CACHE
    if not PLANNING_FACTORS_PATH.exists():
        _PF_CACHE = {}
        return _PF_CACHE
    data = json.loads(PLANNING_FACTORS_PATH.read_text())
    _PF_CACHE = {r["ccn"]: r for r in data.get("ccns", [])}
    return _PF_CACHE


def fc_citation_lookup(ccn: str) -> list[str]:
    """Return a list of human-readable citation lines for a CCN.

    Apex Omega Sec.5.5: time-stamp external data at the point of
    citation. Each line carries (when applicable) source PDF
    filename, printed version date, page numbers, table id, and
    loading driver. CCNs absent from PLANNING_FACTORS.json (e.g.,
    44112 General Warehouse, 45110 Open Storage Area, 61072
    Battalion HQ Admin in CLB-4 because Series 400/500/600 are not
    yet supplied) get a TBD citation that explicitly names the
    missing Series PDF rather than silently inheriting a CLB-4
    observation.
    """
    pf = _load_planning_factors()
    rec = pf.get(ccn)
    if rec is None:
        # Try to identify which Series PDF would carry this CCN.
        # NAVFAC P-72 series mapping: 100s/200s supplied; 300s
        # ranges, 400s storage, 500s medical, 600s admin/community,
        # 700s base support, 800s utilities & ground improvements.
        if not ccn:
            return ["FC 2-000-05N citation: CCN not specified"]
        head = ccn[:1]
        series_for_head = {
            "1": "Series 100 (already supplied)",
            "2": "Series 200 (already supplied)",
            "3": "Series 300 (Training; not yet supplied)",
            "4": "Series 400 (Maintenance & Production; not yet supplied)",
            "5": "Series 500 (Medical; not yet supplied)",
            "6": "Series 600 (Administrative; not yet supplied)",
            "7": "Series 700 (Housing & Community; not yet supplied)",
            "8": "Series 800 (Utilities & Ground; not yet supplied)",
        }
        series_label = series_for_head.get(head, "Series TBD")
        return [
            f"FC 2-000-05N citation: TBD pending {series_label}",
            f"CCN {ccn[:3]} {ccn[3:]} not in supplied Series 100/200 PDFs",
            "Apex Omega rule 4: do not silently inherit CLB-4 values",
        ]
    lines = []
    lines.append(
        f"FC 2-000-05N {rec.get('series', '?')} Series; "
        f"version date {rec.get('source_date', 'unknown')}"
    )
    lines.append(
        f"Source PDF: {rec.get('source_pdf', '?')}"
    )
    pages = rec.get("pages") or []
    if pages:
        lines.append(f"Pages: {pages}")
    tables = rec.get("tables") or []
    if tables:
        for t in tables[:3]:
            tid = t.get("table_id", "?")
            cap = t.get("caption", "")
            ld = t.get("loading_driver", "TBD")
            lines.append(
                f"  {tid} ({cap}); loading driver: {ld}"
            )
            # Track 1c-factor: render the table rows verbatim so
            # Apex Omega rule 7 ("numbers must be traceable") is
            # satisfied at the point of use. Only emit non-empty
            # cells; pdfplumber extraction sometimes pads rows with
            # blank columns where the source table has merged cells.
            for row in (t.get("rows") or [])[:8]:
                non_empty = [str(c).strip() for c in row if c and str(c).strip()]
                if not non_empty:
                    continue
                lines.append(f"    {' | '.join(non_empty)}")
            extra = len(t.get("rows") or []) - 8
            if extra > 0:
                lines.append(f"    ... and {extra} more rows; see "
                             f"audit/PLANNING_FACTORS.yaml CCN {ccn}")
    else:
        # Engineering-study CCN: cite the narrative section count
        narr = rec.get("narrative_sections") or []
        if narr:
            lines.append(
                f"Engineering-study CCN per NAVFAC doctrine; "
                f"{len(narr)} narrative section(s) captured"
            )
            sample = narr[0]
            lines.append(
                f"  e.g. {sample.get('section_id')}: "
                f"{sample.get('first_line', '')[:80]}"
            )
    if rec.get("uom_source"):
        lines.append(f"UoM source: {rec['uom_source']}")
    return lines


def render_fc_citation_footer(ws, ccn: str, start_row: int = 40):
    """Write the FC 2-000-05N citation block as a styled footer on
    a CCN sheet starting at the given row. Uses the Apex Omega
    'calc' palette color (#EAF3F4) and a light border to set the
    block apart from the worksheet body without disturbing the
    cosmetic style of the four CLB-4 reference sheets."""
    lines = fc_citation_lookup(ccn)
    title_cell = ws.cell(row=start_row, column=2,
                         value="FC 2-000-05N Citation (Apex Omega Sec.5.5)")
    title_cell.font = SECTION_FONT
    title_cell.fill = CALC_FILL
    title_cell.border = BOX
    ws.merge_cells(start_row=start_row, start_column=2,
                   end_row=start_row, end_column=8)
    for offset, text in enumerate(lines, start=1):
        body_cell = ws.cell(row=start_row + offset, column=2, value=text)
        body_cell.font = BODY_FONT
        body_cell.fill = CALC_FILL
        body_cell.alignment = LEFT_WRAP
        body_cell.border = BOX
        ws.merge_cells(start_row=start_row + offset, start_column=2,
                       end_row=start_row + offset, end_column=8)
        ws.row_dimensions[start_row + offset].height = 18
    return start_row + len(lines) + 1

INPUT_FILL = PatternFill("solid", fgColor="FFF8DC")
CALC_FILL = PatternFill("solid", fgColor="EAF3F4")
OUTPUT_FILL = PatternFill("solid", fgColor="DCE7C8")
WARNING_FILL = PatternFill("solid", fgColor="F8E2D6")
BANNER_FILL = PatternFill("solid", fgColor="D9D9D9")
SUBHEAD_FILL = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(border_style="thin", color="000000")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_THIN = Border(top=THIN)

TITLE_FONT = Font(name="Calibri", size=16, bold=True)
SECTION_FONT = Font(name="Calibri", size=11, bold=True)
LABEL_FONT_C = Font(name="Calibri", size=11, bold=True)
LABEL_FONT_R = Font(name="Arial", size=10)
COLHEAD_FONT = Font(name="Calibri", size=10, bold=True)
BODY_FONT = Font(name="Calibri", size=10)

CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")

FOOTER_TEXT = "&L\r# DISTRIBUTION: DoD COMMUNITY ONLY"


@dataclass
class UnitProfile:
    uic: str
    tenant_unit: str
    installation: str
    installation_uic: str
    region: str
    planning_area: str
    building_no: str
    planner: str
    fy: str
    project_date: str
    # Track 6: unit_type drives per-unit-type defaults via
    # audit/UNIT_TYPE_DEFAULTS.yaml. Optional; classifier handles
    # absent/unknown values per Apex Omega rule 4.
    unit_type: str = ""

    @classmethod
    def from_json(cls, path: Path) -> "UnitProfile":
        d = json.loads(path.read_text())
        # Drop fields the dataclass does not declare; future profile
        # additions can land in JSON without breaking older code.
        known = {f.name for f in fields(cls)}
        return cls(**{k: v for k, v in d.items() if k in known})


@dataclass
class CcnSpec:
    ccn: str
    loading: float = 0.0
    factor: float = 1.0
    ntg: float = 1.0
    loading_label: str = "Loading driver"
    notes: str = ""
    facility_name: Optional[str] = None
    uom: Optional[str] = None
    pattern: str = "default"
    pattern_data: dict = field(default_factory=dict)


def load_vocabulary() -> dict:
    if not VOCAB_PATH.exists():
        sys.exit(f"FATAL: missing {VOCAB_PATH}")
    data = json.loads(VOCAB_PATH.read_text())
    return {v["ccn"]: v for v in data["ccns"]}


def vocab_uom(rec: dict) -> str:
    for key in ("uom_area", "uom_other", "uom_alt"):
        v = rec.get(key)
        if v:
            return v.strip("[]")
    return "EA"


def gsf_or_gsy(uom: str) -> str:
    u = (uom or "").upper()
    if u in ("SF", "GSF"):
        return "GSF"
    if u in ("SY", "GSY"):
        return "GSY"
    if u in ("AC",):
        return "AC"
    if u in ("EA", "LF", "LS"):
        return u
    return u or "GSF"


def apply_page_setup(ws):
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.scale = 84
    ws.page_margins = PageMargins(left=0.75, right=0.5, top=0.75,
                                  bottom=0.5, header=0.0, footer=0.3)
    ws.print_options = PrintOptions(horizontalCentered=False,
                                    verticalCentered=False)
    ws.oddFooter.left.text = "\r# DISTRIBUTION: DoD COMMUNITY ONLY"


def set_col_widths(ws):
    ws.column_dimensions["A"].width = 2.8164
    for col in range(2, 33):
        ws.column_dimensions[get_column_letter(col)].width = 8.0
    ws.column_dimensions["AG"].hidden = True
    ws.column_dimensions["AH"].hidden = True
    ws.column_dimensions["AG"].width = 9.18
    ws.column_dimensions["AH"].width = 9.18


def set_row_heights(ws, total_data_row=37):
    heights = {1: 15.75, 2: 15.75, 3: 30.75, 4: 14.5, 5: 15.75,
               6: 28.5, 7: 14.5}
    for r, h in heights.items():
        ws.row_dimensions[r].height = h
    for r in range(8, total_data_row):
        ws.row_dimensions[r].height = 15.75
    for r in (8, 9, 10, 11, 12, 13):
        ws.row_dimensions[r].height = 12.75
    ws.row_dimensions[total_data_row].height = 12.75


def merge_and_set(ws, rng: str, value=None, font=None, fill=None,
                  alignment=None, border=None):
    ws.merge_cells(rng)
    top_left = rng.split(":")[0]
    cell = ws[top_left]
    if value is not None:
        cell.value = value
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def banner_block(ws, profile: UnitProfile, ccn_display: str,
                 facility_name: str):
    merge_and_set(ws, "A1:AF2",
                  value="BASIC FACILITY REQUIREMENTS WORKSHEET",
                  font=TITLE_FONT, alignment=CENTER)
    merge_and_set(ws, "A3:E3", value="Installation:",
                  font=LABEL_FONT_C, alignment=LEFT_WRAP)
    merge_and_set(ws, "F3:L3", value=profile.installation,
                  font=LABEL_FONT_R, alignment=LEFT_WRAP)
    merge_and_set(ws, "M3:Q3", value="Tenant:",
                  font=LABEL_FONT_C, alignment=LEFT_WRAP)
    merge_and_set(ws, "R3:AC3", value=profile.tenant_unit,
                  font=LABEL_FONT_R, alignment=LEFT_WRAP)
    merge_and_set(ws, "AE3:AF3")
    merge_and_set(ws, "A4:E4", value="Installation UIC:",
                  font=LABEL_FONT_C, alignment=LEFT_WRAP)
    merge_and_set(ws, "F4:L4", value=profile.installation_uic,
                  font=LABEL_FONT_R, alignment=LEFT_WRAP)
    merge_and_set(ws, "M4:Q4", value="Tenant UIC:",
                  font=LABEL_FONT_C, alignment=LEFT_WRAP)
    merge_and_set(ws, "R4:X4", value=profile.uic,
                  font=LABEL_FONT_R, alignment=LEFT_WRAP)
    merge_and_set(ws, "Y4:AF4",
                  value=f"Planning Area: {profile.planning_area}",
                  font=LABEL_FONT_C, alignment=LEFT_WRAP)
    merge_and_set(ws, "A5:B5", value="CCN:",
                  font=LABEL_FONT_C, alignment=LEFT_WRAP)
    merge_and_set(ws, "C5:L5", value=ccn_display,
                  font=LABEL_FONT_R, alignment=LEFT_WRAP)
    merge_and_set(ws, "M5:Q5", value="Nomenclature:",
                  font=LABEL_FONT_C, alignment=LEFT_WRAP)
    merge_and_set(ws, "R5:AF5", value=facility_name,
                  font=LABEL_FONT_R, alignment=LEFT_WRAP)
    section_text = (f"SPACE REQUIREMENT ANALYSIS  ,  CCN {ccn_display}  ,  "
                    f"FC 2-000-05N (Series 100, 11 Feb 2026)")
    merge_and_set(ws, "A6:AF6", value=section_text,
                  font=SECTION_FONT, fill=BANNER_FILL,
                  alignment=Alignment(horizontal="left", vertical="center",
                                      wrap_text=True))


def column_headers(ws, label_a7: str = "", header_v7: str = "",
                   header_y7: str = "", header_ab7: str = "",
                   row: int = 7,
                   label_a: Optional[str] = None,
                   header_v: Optional[str] = None,
                   header_y: Optional[str] = None,
                   header_ab: Optional[str] = None):
    """Place a four-block column header row (A:U, V:X, Y:AA, AB:AF).

    Backwards compatibility: callers may pass label_a7/header_v7/header_y7/
    header_ab7 (default row=7). New callers can pass row=N and the same
    label arguments without the "7" suffix.
    """
    la = label_a if label_a is not None else label_a7
    hv = header_v if header_v is not None else header_v7
    hy = header_y if header_y is not None else header_y7
    hab = header_ab if header_ab is not None else header_ab7
    merge_and_set(ws, f"A{row}:U{row}", value=la,
                  font=SECTION_FONT, fill=SUBHEAD_FILL,
                  alignment=LEFT_WRAP, border=BOX)
    merge_and_set(ws, f"V{row}:X{row}", value=hv,
                  font=COLHEAD_FONT, fill=SUBHEAD_FILL,
                  alignment=CENTER_WRAP, border=BOX)
    merge_and_set(ws, f"Y{row}:AA{row}", value=hy,
                  font=COLHEAD_FONT, fill=SUBHEAD_FILL,
                  alignment=CENTER_WRAP, border=BOX)
    merge_and_set(ws, f"AB{row}:AF{row}", value=hab,
                  font=COLHEAD_FONT, fill=SUBHEAD_FILL,
                  alignment=CENTER_WRAP, border=BOX)


def data_row(ws, row: int, label: str, factor_val, count_val,
             total_formula: str, count_role: str = "input"):
    merge_and_set(ws, f"A{row}:U{row}", value=label,
                  font=BODY_FONT, alignment=LEFT_WRAP, border=BOX)
    merge_and_set(ws, f"V{row}:X{row}", value=factor_val,
                  font=BODY_FONT, alignment=CENTER, border=BOX,
                  fill=INPUT_FILL)
    merge_and_set(ws, f"Y{row}:AA{row}", value=count_val,
                  font=BODY_FONT, alignment=CENTER, border=BOX,
                  fill=(INPUT_FILL if count_role == "input" else CALC_FILL))
    merge_and_set(ws, f"AB{row}:AF{row}", value=total_formula,
                  font=BODY_FONT, alignment=CENTER, border=BOX,
                  fill=CALC_FILL)


def subtotal_row(ws, row: int, label: str, formula: str):
    merge_and_set(ws, f"A{row}:AA{row}", value=label,
                  font=LABEL_FONT_C, alignment=LEFT_WRAP, border=BOX)
    merge_and_set(ws, f"AB{row}:AF{row}", value=formula,
                  font=BODY_FONT, alignment=CENTER, border=BOX,
                  fill=CALC_FILL)


def total_requirement_row(ws, row: int, value_formula: str, uom_label: str):
    merge_and_set(ws, f"A{row}:G{row}", value="TOTAL REQUIREMENT:",
                  font=LABEL_FONT_C, alignment=Alignment(horizontal="right",
                                                         vertical="center"),
                  border=TOP_THIN)
    merge_and_set(ws, f"H{row}:K{row}", value=value_formula,
                  font=BODY_FONT, alignment=CENTER_WRAP,
                  fill=OUTPUT_FILL, border=BOX)
    fmt = '#,##0\\ "' + gsf_or_gsy(uom_label) + '"'
    ws[f"H{row}"].number_format = fmt
    merge_and_set(ws, f"O{row}:Q{row}", value=gsf_or_gsy(uom_label),
                  font=LABEL_FONT_C, alignment=CENTER, border=TOP_THIN)
    merge_and_set(ws, f"AB{row}:AF{row}", value="",
                  font=LABEL_FONT_C, alignment=Alignment(horizontal="right",
                                                         vertical="center"))


def section_banner(ws, row: int, text: str):
    merge_and_set(ws, f"A{row}:AF{row}", value=text,
                  font=SECTION_FONT, fill=BANNER_FILL,
                  alignment=Alignment(horizontal="left", vertical="center",
                                      wrap_text=True))


def common_setup(ws, profile: UnitProfile, spec: CcnSpec):
    ccn_display = f"{spec.ccn[:3]} {spec.ccn[3:]}"
    set_col_widths(ws)
    set_row_heights(ws)
    apply_page_setup(ws)
    ws.print_area = "A1:AF37"
    banner_block(ws, profile, ccn_display, spec.facility_name or "")
    return ccn_display


def make_default_sheet(wb: Workbook, profile: UnitProfile, spec: CcnSpec):
    ws = wb.create_sheet(spec.ccn)
    common_setup(ws, profile, spec)
    column_headers(
        ws,
        label_a7=f"{spec.facility_name or spec.ccn} (Loading driver: "
                 f"{spec.loading_label})",
        header_v7="Factor",
        header_y7="Count",
        header_ab7=gsf_or_gsy(spec.uom or ""),
    )
    data_row(ws, 8, label="Primary loading row",
             factor_val=spec.factor, count_val=spec.loading,
             total_formula=f"=V8*Y8*{spec.ntg}")
    for r in range(9, 14):
        data_row(ws, r, label="", factor_val="", count_val="",
                 total_formula=f"=V{r}*Y{r}*{spec.ntg}")
    subtotal_row(ws, 15, "Subtotal NSF", "=SUM(AB8:AF13)")
    subtotal_row(ws, 16, f"Net to Gross multiplier (NTG = {spec.ntg})",
                 "=AB15*1")
    total_requirement_row(ws, 37, value_formula="=AB16",
                          uom_label=spec.uom or "")
    return ws


def make_primary_items_sheet(wb: Workbook, profile: UnitProfile,
                             spec: CcnSpec):
    """Pattern: primary_items.

    Shape observed in one example unit (CLB-4 14345 Armory). Not a
    canonical Layer 5 pattern; ratify against FC 2-000-05N before
    relying on it for any new unit type. Different facility CCNs may
    use entirely different row shapes.

    pattern_data:
      row7_label : str (e.g. "Sensitive Item Category")
      v_label    : str (e.g. "GSF / Item")
      y_label    : str (e.g. "Count")
      ab_label   : str (e.g. "GSF")
      items      : list of {label, sf_per, count}, up to 6 entries
      subtotal_label : str (e.g. "Total Sensitive Items | Total NSF")
      roundup_label  : str (e.g. "Required GSF = ROUNDUP(...)")
    """
    pd = spec.pattern_data
    ws = wb.create_sheet(spec.ccn)
    common_setup(ws, profile, spec)
    column_headers(ws,
                   label_a7=pd.get("row7_label", "Item Category"),
                   header_v7=pd.get("v_label", "GSF / Item"),
                   header_y7=pd.get("y_label", "Count"),
                   header_ab7=pd.get("ab_label",
                                     gsf_or_gsy(spec.uom or "")))
    items = pd.get("items", [])[:6]
    for i, it in enumerate(items):
        r = 8 + i
        data_row(ws, r,
                 label=it.get("label", ""),
                 factor_val=it.get("sf_per", ""),
                 count_val=it.get("count", ""),
                 total_formula=f"=V{r}*Y{r}")
    for r in range(8 + len(items), 14):
        data_row(ws, r, label="", factor_val="", count_val="",
                 total_formula=f"=V{r}*Y{r}")
    subtotal_row(ws, 15,
                 pd.get("subtotal_label", "Subtotal NSF"),
                 "=SUM(AB8:AB13)")
    subtotal_row(ws, 16,
                 pd.get("roundup_label", "Required GSF = ROUNDUP(NSF, 0)"),
                 "=ROUNDUP(AB15,0)")
    total_requirement_row(ws, 37, value_formula="=AB16",
                          uom_label=spec.uom or "")
    return ws


def make_admin_sheet(wb: Workbook, profile: UnitProfile, spec: CcnSpec):
    """Pattern: admin.

    Shape observed in one example unit (CLB-4 61072 BN HQ Admin).
    Not a canonical Layer 5 pattern; ratify against FC 2-000-05N
    Series 100 admin facility tables before relying on it. SF/person
    defaults (120 for officers, 60 for enlisted cubicles) are
    extracted from CLB-4 SW and must be checked against the current
    FC 2-000-05N edition for the unit type at hand.

    pattern_data:
      excluded_billets : list of strings (rows 8 plus, optional)
      officers   : {label, sf_per, count}
      enlisted   : {label, sf_per, count}
      ntg        : float
    """
    pd = spec.pattern_data
    ws = wb.create_sheet(spec.ccn)
    common_setup(ws, profile, spec)

    excluded = pd.get("excluded_billets") or []
    if excluded:
        section_banner(ws, 7,
                       "BILLETS EXCLUDED FROM THIS CCN "
                       "(require separate facility CCNs)")
        for i, line in enumerate(excluded[:5]):
            r = 8 + i
            merge_and_set(ws, f"A{r}:AF{r}", value=line, font=BODY_FONT,
                          alignment=LEFT_WRAP, border=BOX)
        calc_banner_row = 13
    else:
        calc_banner_row = 7

    section_banner(ws, calc_banner_row, "ADMINISTRATIVE SPACE CALCULATION")
    hdr = calc_banner_row + 1
    column_headers(ws, row=hdr,
                   label_a=pd.get("calc_label", "Space Type"),
                   header_v="SF / Person",
                   header_y="Count",
                   header_ab="NSF")
    officers = pd.get("officers") or {"label": "Private Offices",
                                       "sf_per": 120, "count": 0}
    enlisted = pd.get("enlisted") or {"label": "Cubicles",
                                       "sf_per": 60, "count": 0}
    r_off = hdr + 1
    r_enl = hdr + 2
    data_row(ws, r_off, label=officers.get("label", "Private Offices"),
             factor_val=officers.get("sf_per", 120),
             count_val=officers.get("count", 0),
             total_formula=f"=V{r_off}*Y{r_off}")
    data_row(ws, r_enl, label=enlisted.get("label", "Cubicles"),
             factor_val=enlisted.get("sf_per", 60),
             count_val=enlisted.get("count", 0),
             total_formula=f"=V{r_enl}*Y{r_enl}")
    r_sub = hdr + 3
    subtotal_row(ws, r_sub, "Total NSF",
                 f"=SUM(AB{r_off}:AB{r_enl})")
    ntg = pd.get("ntg", 1.33)
    r_total_banner = r_sub + 2
    section_banner(ws, r_total_banner, "TOTAL SPACE REQUIREMENT")
    r_round = r_total_banner + 1
    merge_and_set(ws, f"A{r_round}:U{r_round}",
                  value=f"NTG Factor = {ntg}, Required GSF = "
                        f"ROUNDUP(NSF * NTG, 0)",
                  font=BODY_FONT, alignment=LEFT_WRAP, border=BOX)
    merge_and_set(ws, f"V{r_round}:X{r_round}", value=ntg,
                  font=BODY_FONT, alignment=CENTER, border=BOX,
                  fill=INPUT_FILL)
    merge_and_set(ws, f"Y{r_round}:AA{r_round}", value="",
                  font=BODY_FONT, alignment=CENTER, border=BOX)
    merge_and_set(ws, f"AB{r_round}:AF{r_round}",
                  value=f"=ROUNDUP(AB{r_sub}*V{r_round},0)",
                  font=BODY_FONT, alignment=CENTER, border=BOX,
                  fill=CALC_FILL)
    total_requirement_row(ws, 37, value_formula=f"=AB{r_round}",
                          uom_label=spec.uom or "")
    return ws


def make_shop_with_bays_sheet(wb: Workbook, profile: UnitProfile,
                              spec: CcnSpec):
    """Pattern: shop_with_bays.

    Shape observed in one example unit (CLB-4 21451 Auto Org Shop).
    Not a canonical Layer 5 pattern; ratify against FC 2-000-05N
    Series 200 maintenance facility tables before relying on it.
    SF/bay (420) and bays-per-N-vehicles (30) numbers come from the
    CLB-4 example and must be verified against the current
    FC 2-000-05N edition for the maintenance facility type and unit
    in question. Aviation maintenance shops, ammo shops, comm shops,
    EOD shops, and field maintenance shops likely have different
    factors.

    pattern_data:
      officers   : {label, sf_per, count}
      snco       : {label, sf_per, count}
      vehicles   : list of {label, qty}, up to 3 entries
      bays_per_n_vehicles : int
      sf_per_bay : float
      ntg        : float
    """
    pd = spec.pattern_data
    ws = wb.create_sheet(spec.ccn)
    common_setup(ws, profile, spec)

    section_banner(ws, 7, "SECTION 1 - ADMINISTRATIVE SPACE")
    merge_and_set(ws, "A8:U8", value="Personnel Category",
                  font=COLHEAD_FONT, fill=SUBHEAD_FILL,
                  alignment=LEFT_WRAP, border=BOX)
    merge_and_set(ws, "V8:X8", value="SF / Person",
                  font=COLHEAD_FONT, fill=SUBHEAD_FILL,
                  alignment=CENTER_WRAP, border=BOX)
    merge_and_set(ws, "Y8:AA8", value="Count",
                  font=COLHEAD_FONT, fill=SUBHEAD_FILL,
                  alignment=CENTER_WRAP, border=BOX)
    merge_and_set(ws, "AB8:AF8", value="NSF",
                  font=COLHEAD_FONT, fill=SUBHEAD_FILL,
                  alignment=CENTER_WRAP, border=BOX)
    officers = pd.get("officers") or {"label": "Officers - Private Offices",
                                       "sf_per": 120, "count": 0}
    snco = pd.get("snco") or {"label": "SNCO/Staff - Cubicles",
                              "sf_per": 60, "count": 0}
    data_row(ws, 9, label=officers.get("label", "Officers"),
             factor_val=officers.get("sf_per", 120),
             count_val=officers.get("count", 0),
             total_formula="=V9*Y9")
    data_row(ws, 10, label=snco.get("label", "SNCO/Staff"),
             factor_val=snco.get("sf_per", 60),
             count_val=snco.get("count", 0),
             total_formula="=V10*Y10")
    subtotal_row(ws, 11, "Administrative Subtotal", "=SUM(AB9:AB10)")

    section_banner(ws, 13, "SECTION 2 - MAINTENANCE BAYS")
    merge_and_set(ws, "A14:U14", value="Vehicle Category",
                  font=COLHEAD_FONT, fill=SUBHEAD_FILL,
                  alignment=LEFT_WRAP, border=BOX)
    merge_and_set(ws, "V14:X14", value="SF / Bay",
                  font=COLHEAD_FONT, fill=SUBHEAD_FILL,
                  alignment=CENTER_WRAP, border=BOX)
    merge_and_set(ws, "Y14:AA14", value="Qty",
                  font=COLHEAD_FONT, fill=SUBHEAD_FILL,
                  alignment=CENTER_WRAP, border=BOX)
    merge_and_set(ws, "AB14:AF14", value="NSF",
                  font=COLHEAD_FONT, fill=SUBHEAD_FILL,
                  alignment=CENTER_WRAP, border=BOX)
    sf_per_bay = pd.get("sf_per_bay", 420)
    vehicles = pd.get("vehicles", [])[:3]
    for i, v in enumerate(vehicles):
        r = 15 + i
        merge_and_set(ws, f"A{r}:U{r}",
                      value=v.get("label", ""),
                      font=BODY_FONT, alignment=LEFT_WRAP, border=BOX)
        merge_and_set(ws, f"V{r}:X{r}", value=sf_per_bay,
                      font=BODY_FONT, alignment=CENTER, border=BOX)
        merge_and_set(ws, f"Y{r}:AA{r}", value=v.get("qty", 0),
                      font=BODY_FONT, alignment=CENTER, border=BOX,
                      fill=INPUT_FILL)
        merge_and_set(ws, f"AB{r}:AF{r}", value="",
                      font=BODY_FONT, alignment=CENTER, border=BOX)
    for i in range(len(vehicles), 3):
        r = 15 + i
        merge_and_set(ws, f"A{r}:U{r}", value="",
                      font=BODY_FONT, alignment=LEFT_WRAP, border=BOX)
        merge_and_set(ws, f"V{r}:X{r}", value="",
                      font=BODY_FONT, alignment=CENTER, border=BOX)
        merge_and_set(ws, f"Y{r}:AA{r}", value=0,
                      font=BODY_FONT, alignment=CENTER, border=BOX,
                      fill=INPUT_FILL)
        merge_and_set(ws, f"AB{r}:AF{r}", value="",
                      font=BODY_FONT, alignment=CENTER, border=BOX)
    bays_n = pd.get("bays_per_n_vehicles", 30)
    merge_and_set(ws, "A18:U18",
                  value=f"Total Self-Propelled Vehicles | "
                        f"CEILING(total / {bays_n}) bays",
                  font=BODY_FONT, alignment=LEFT_WRAP, border=BOX)
    merge_and_set(ws, "V18:X18", value="",
                  font=BODY_FONT, alignment=CENTER, border=BOX)
    merge_and_set(ws, "Y18:AA18", value="=SUM(Y15:Y17)",
                  font=BODY_FONT, alignment=CENTER, border=BOX,
                  fill=CALC_FILL)
    merge_and_set(ws, "AB18:AF18", value="",
                  font=BODY_FONT, alignment=CENTER, border=BOX)
    merge_and_set(ws, "A19:U19",
                  value=f"Maintenance Bays (CEILING bays per "
                        f"{bays_n} vehicles)",
                  font=BODY_FONT, alignment=LEFT_WRAP, border=BOX)
    merge_and_set(ws, "V19:X19", value=sf_per_bay,
                  font=BODY_FONT, alignment=CENTER, border=BOX)
    merge_and_set(ws, "Y19:AA19", value=f"=CEILING(Y18/{bays_n},1)",
                  font=BODY_FONT, alignment=CENTER, border=BOX,
                  fill=CALC_FILL)
    merge_and_set(ws, "AB19:AF19", value="=V19*Y19",
                  font=BODY_FONT, alignment=CENTER, border=BOX,
                  fill=CALC_FILL)
    subtotal_row(ws, 20, "Maintenance Bays Subtotal", "=AB19")

    section_banner(ws, 22, "TOTAL SPACE REQUIREMENT")
    subtotal_row(ws, 23, "Total NSF (Administrative + Maintenance Bays)",
                 "=AB11+AB20")
    ntg = pd.get("ntg", 1.33)
    merge_and_set(ws, "A24:U24",
                  value=f"NTG Factor = {ntg}, Required GSF = "
                        f"ROUNDUP(NSF * NTG, 0)",
                  font=BODY_FONT, alignment=LEFT_WRAP, border=BOX)
    merge_and_set(ws, "V24:X24", value=ntg,
                  font=BODY_FONT, alignment=CENTER, border=BOX,
                  fill=INPUT_FILL)
    merge_and_set(ws, "Y24:AA24", value="",
                  font=BODY_FONT, alignment=CENTER, border=BOX)
    merge_and_set(ws, "AB24:AF24",
                  value=f"=ROUNDUP(AB23*V24,0)",
                  font=BODY_FONT, alignment=CENTER, border=BOX,
                  fill=CALC_FILL)
    total_requirement_row(ws, 37, value_formula="=AB24",
                          uom_label=spec.uom or "")
    return ws


PATTERN_DISPATCH = {
    "default": make_default_sheet,
    "primary_items": make_primary_items_sheet,
    "admin": make_admin_sheet,
    "shop_with_bays": make_shop_with_bays_sheet,
}


def make_ccn_sheet(wb: Workbook, profile: UnitProfile, spec: CcnSpec):
    handler = PATTERN_DISPATCH.get(spec.pattern, make_default_sheet)
    ws = handler(wb, profile, spec)
    # Apex Omega Sec.5.5: render FC 2-000-05N citation footer on every
    # CCN sheet so the source of the planning factor is visible inline
    # on the BFR. Citations are read from audit/PLANNING_FACTORS.yaml
    # via fc_citation_lookup. CCNs absent from supplied Series 100/200
    # carry an explicit "TBD pending Series N" citation rather than
    # silently inheriting CLB-4 values.
    if ws is not None:
        render_fc_citation_footer(ws, spec.ccn, start_row=40)
    return ws


def make_to_sheet(wb: Workbook, profile: UnitProfile):
    ws = wb.create_sheet("TO")
    apply_page_setup(ws)
    ws.row_dimensions[6].height = 30.0
    headers = ["", "LINE", "NOTE", "CCN", "CCN Description", "UIC",
               "Rec CD", "BIC", "Billet Description", "Alpha Grade",
               "BMOS", "PMOS"]
    for c, val in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=c, value=val)
        cell.font = COLHEAD_FONT
        cell.alignment = CENTER_WRAP
        cell.fill = SUBHEAD_FILL
        cell.border = BOX
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 7
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 11
    ws.column_dimensions["K"].width = 7
    ws.column_dimensions["L"].width = 7
    return ws


def make_te_sheet(wb: Workbook, profile: UnitProfile):
    ws = wb.create_sheet("TE")
    apply_page_setup(ws)
    ws.row_dimensions[4].height = 30.0
    headers = ["", "ROW", "NOTE", "CCN", "CCN Description",
               "CCN 21451 Equipment Type", "UIC", "TAMCN Short", "TAMCN",
               "Nomenclature", "TAM Stat", "U/I", "Rdy", "Ind Qty",
               "Org Qty", "Unit T/E", "L, Ft", "W, Ft", "H, Ft",
               "Volume Ea", "Volume Total"]
    for c, val in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=val)
        cell.font = COLHEAD_FONT
        cell.alignment = CENTER_WRAP
        cell.fill = SUBHEAD_FILL
        cell.border = BOX
    return ws


def make_unit_rollup(wb: Workbook, profile: UnitProfile,
                     ccns: List[CcnSpec]):
    ws = wb.create_sheet("UNIT_ROLLUP", 0)
    set_col_widths(ws)
    apply_page_setup(ws)
    merge_and_set(ws, "A1:AF2",
                  value="BASIC FACILITY REQUIREMENTS UNIT ROLLUP",
                  font=TITLE_FONT, alignment=CENTER)
    merge_and_set(ws, "A3:E3", value="Installation:",
                  font=LABEL_FONT_C, alignment=LEFT_WRAP)
    merge_and_set(ws, "F3:L3", value=profile.installation,
                  font=LABEL_FONT_R, alignment=LEFT_WRAP)
    merge_and_set(ws, "M3:Q3", value="Tenant:",
                  font=LABEL_FONT_C, alignment=LEFT_WRAP)
    merge_and_set(ws, "R3:AC3", value=profile.tenant_unit,
                  font=LABEL_FONT_R, alignment=LEFT_WRAP)
    merge_and_set(ws, "A4:E4", value="Installation UIC:",
                  font=LABEL_FONT_C, alignment=LEFT_WRAP)
    merge_and_set(ws, "F4:L4", value=profile.installation_uic,
                  font=LABEL_FONT_R, alignment=LEFT_WRAP)
    merge_and_set(ws, "M4:Q4", value="Tenant UIC:",
                  font=LABEL_FONT_C, alignment=LEFT_WRAP)
    merge_and_set(ws, "R4:X4", value=profile.uic,
                  font=LABEL_FONT_R, alignment=LEFT_WRAP)
    merge_and_set(ws, "Y4:AF4",
                  value=f"Planning Area: {profile.planning_area}",
                  font=LABEL_FONT_C, alignment=LEFT_WRAP)

    headers = ["CCN", "Description", "UM", "Total"]
    base = 8
    for c, val in enumerate(headers, start=2):
        cell = ws.cell(row=base, column=c, value=val)
        cell.font = COLHEAD_FONT
        cell.alignment = CENTER_WRAP
        cell.fill = SUBHEAD_FILL
        cell.border = BOX
    for i, spec in enumerate(ccns):
        r = base + 1 + i
        ccn_display = f"{spec.ccn[:3]} {spec.ccn[3:]}"
        ws.cell(row=r, column=2, value=ccn_display).font = BODY_FONT
        ws.cell(row=r, column=3, value=spec.facility_name).font = BODY_FONT
        ws.cell(row=r, column=4, value=gsf_or_gsy(spec.uom or "")).font = BODY_FONT
        ws.cell(row=r, column=5, value=f"='{spec.ccn}'!H37").font = BODY_FONT
        ws.cell(row=r, column=5).fill = OUTPUT_FILL
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = BOX
            ws.cell(row=r, column=c).alignment = CENTER_WRAP
    total_row = base + 1 + len(ccns) + 1
    ws.cell(row=total_row, column=2, value="GRAND TOTAL").font = LABEL_FONT_C
    ws.cell(row=total_row, column=5,
            value=f"=SUM(E{base+1}:E{base+len(ccns)})").font = LABEL_FONT_C
    ws.cell(row=total_row, column=5).fill = OUTPUT_FILL
    for c in range(2, 6):
        ws.cell(row=total_row, column=c).border = TOP_THIN
    return ws


def build(profile: UnitProfile, ccns: List[CcnSpec]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    make_unit_rollup(wb, profile, ccns)
    make_to_sheet(wb, profile)
    make_te_sheet(wb, profile)
    for spec in ccns:
        make_ccn_sheet(wb, profile, spec)
    wb.calculation.fullCalcOnLoad = True
    return wb


def hydrate_ccns(raw_specs: list, vocab: dict) -> List[CcnSpec]:
    out = []
    for s in raw_specs:
        ccn = s["ccn"]
        v = vocab.get(ccn, {})
        spec = CcnSpec(
            ccn=ccn,
            loading=float(s.get("loading", 0)),
            factor=float(s.get("factor", 1.0)),
            ntg=float(s.get("ntg", 1.0)),
            loading_label=s.get("loading_label", "Loading driver"),
            notes=s.get("notes", ""),
            facility_name=s.get("facility_name") or v.get("title") or ccn,
            uom=s.get("uom") or vocab_uom(v),
            pattern=s.get("pattern", "default"),
            pattern_data=s.get("pattern_data", {}),
        )
        out.append(spec)
    return out


def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--profile", type=Path, required=True)
    ap.add_argument("--ccns", type=Path, required=True)
    ap.add_argument("--output", type=Path, required=True)
    args = ap.parse_args()

    profile = UnitProfile.from_json(args.profile)
    raw = json.loads(args.ccns.read_text())
    vocab = load_vocabulary()
    ccns = hydrate_ccns(raw, vocab)

    wb = build(profile, ccns)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Wrote {args.output}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
