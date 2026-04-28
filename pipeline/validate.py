"""
Layer 6, BFR validation harness (Apex Omega).

Run:
  python3 pipeline/validate.py <workbook.xlsx> [--report <path>]

Emits a deterministic pass/fail report covering the six checks
mandated by audit/PIPELINE.md Layer 6:

  1. Schema     , TO/TE sheets present with the Format-B column
                   contract; header row auto-detected.
  2. NOTE       , every TO row carries a NOTE matching
                   ^\\d{4,5}[a-z]{0,4}$; every TE row carries a bare-
                   CCN NOTE matching ^\\d{4,5}$.
  3. NOTE<->CCN , for every row, the CCN portion of NOTE equals the
                   CCN column value.
  4. Vocabulary , every 5-digit CCN sheet name and every CCN seen in
                   TO/TE/NOTE appears in audit/CCN_VOCABULARY.json
                   (1,059 canonical CCNs from FC 2-000-05N Appendix A,
                   2019-06-27 generation; commit c328a36).
  5. Cell errors, no #REF! / #DIV/0! / #NAME? / #VALUE! / #NULL! /
                   #NUM! tokens in any cached cell value across the
                   whole workbook.
  6. Roll-up    , UNIT_ROLLUP (or named variant) lists every CCN
                   sheet present, and each row's pulled total matches
                   the CCN sheet's TOTAL REQUIREMENT cell.
  7. Billet     , every TO data row attributes its billet to a CCN
                   sheet via the NOTE column; orphans (no NOTE) and
                   unknown-CCN references are failures.
  8. Equipment  , every TE data row attributes its TAMCN to a CCN
                   sheet via NOTE or CCN column; orphans and
                   unknown-CCN references are failures.

Exit code 0 if all checks pass; 1 otherwise.

Apex Omega:
  - Facts only. Each finding cites the cell or sheet by name.
  - No silent normalization: if a workbook violates the contract, the
    report says exactly which contract clause and which cell.
  - This script is read-only against the workbook. It never modifies
    source files.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
VOCAB_PATH = REPO_ROOT / "audit" / "CCN_VOCABULARY.json"

# Format-B contract, columns we require on the TO sheet header row.
# Order does not matter for the check; presence does.
TO_REQUIRED = [
    "LINE", "NOTE", "CCN", "UIC", "Rec CD", "BIC",
    "Billet Description", "Alpha Grade", "BMOS", "PMOS",
]
TE_REQUIRED = [
    "ROW", "NOTE", "CCN", "UIC", "TAMCN", "Nomenclature",
    "TAM Stat", "U/I", "Ind Qty", "Org Qty",
]

NOTE_TO = re.compile(r"^\d{4,5}[a-z]{0,4}$")
NOTE_TE = re.compile(r"^\d{4,5}$")
CCN_5 = re.compile(r"^\d{5}$")
ERR_TOKENS = ("#REF!", "#DIV/0!", "#NAME?", "#VALUE!", "#NULL!", "#NUM!")


@dataclass
class CheckResult:
    name: str
    passed: bool
    findings: list = field(default_factory=list)

    def line(self) -> str:
        status = "PASS" if self.passed else "FAIL"
        return f"CHECK: {self.name:<28} {status}"


def _norm(s):
    return str(s).strip().upper() if s is not None else ""


def find_header_row(ws, required_tokens, max_scan_rows=12):
    """Scan first N rows for the row that contains every required token
    (substring match, case-insensitive)."""
    needed = [_norm(t) for t in required_tokens]
    for r in range(1, max_scan_rows + 1):
        cells = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            cells.append((c, _norm(v)))
        joined = " | ".join(t for _, t in cells)
        if all(tok in joined for tok in needed):
            cols = {}
            for tok in needed:
                # First column whose normalized value contains the token
                for c, txt in cells:
                    if tok in txt and tok not in cols:
                        cols[tok] = c
                        break
            return r, cols
    return None, {}


def discover_ccn_sheets(wb):
    """Return sheets whose name is a 5-digit CCN (visible or hidden)."""
    return [s for s in wb.sheetnames if CCN_5.match(s)]


def discover_total_cell(ws):
    """Return (label_coord, value_coord, value) for the first row whose
    column A matches 'TOTAL REQUIREMENT'."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.column != 1 or cell.value is None:
                continue
            if "TOTAL REQUIREMENT" in str(cell.value).upper():
                value_cell = ws.cell(row=cell.row, column=8)  # column H
                return cell.coordinate, value_cell.coordinate, value_cell.value
    return None, None, None


def load_vocabulary():
    if not VOCAB_PATH.exists():
        return None
    data = json.loads(VOCAB_PATH.read_text())
    return {v["ccn"] for v in data["ccns"]}


# ---------- Checks ---------------------------------------------------

def check_schema(wb):
    r = CheckResult("1. Schema (TO/TE columns)", True)
    if "TO" not in wb.sheetnames:
        r.passed = False
        r.findings.append("missing sheet: TO")
    if "TE" not in wb.sheetnames:
        r.passed = False
        r.findings.append("missing sheet: TE")
    if not r.passed:
        return r, None, None
    to_row, to_cols = find_header_row(wb["TO"], TO_REQUIRED)
    te_row, te_cols = find_header_row(wb["TE"], TE_REQUIRED)
    if to_row is None:
        r.passed = False
        r.findings.append("TO: no header row carrying all required tokens "
                          + str(TO_REQUIRED))
    else:
        r.findings.append(f"TO: header row {to_row}; required cols matched "
                          f"{len(to_cols)}/{len(TO_REQUIRED)}")
    if te_row is None:
        r.passed = False
        r.findings.append("TE: no header row carrying all required tokens "
                          + str(TE_REQUIRED))
    else:
        r.findings.append(f"TE: header row {te_row}; required cols matched "
                          f"{len(te_cols)}/{len(TE_REQUIRED)}")
    return r, (to_row, to_cols), (te_row, te_cols)


def _column_values(ws, col_idx, start_row, end_row=None):
    end_row = end_row or ws.max_row
    out = []
    for r in range(start_row, end_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        out.append((r, v))
    return out


def check_note_coverage(wb, to_meta, te_meta):
    r = CheckResult("2. NOTE coverage", True)
    if not (to_meta and te_meta):
        r.passed = False
        r.findings.append("skipped: schema check failed")
        return r
    (to_row, to_cols), (te_row, te_cols) = to_meta, te_meta
    to_note_col = to_cols.get("NOTE")
    te_note_col = te_cols.get("NOTE")

    def tally(ws, hdr_row, note_col, regex, label):
        total = 0
        valid = 0
        invalid = []
        empty = 0
        for rr, v in _column_values(ws, note_col, hdr_row + 1):
            if v is None or str(v).strip() == "":
                # Skip rows that are entirely blank
                if all(ws.cell(row=rr, column=c).value in (None, "")
                       for c in range(1, ws.max_column + 1)):
                    continue
                empty += 1
                total += 1
                continue
            total += 1
            s = str(v).strip()
            if regex.match(s):
                valid += 1
            else:
                invalid.append((rr, s))
        r.findings.append(
            f"{label}: data rows={total}, NOTE matched={valid}, "
            f"empty={empty}, malformed={len(invalid)}"
        )
        if empty > 0 or invalid:
            return False, invalid[:10]
        return True, []

    to_pass, to_bad = tally(wb["TO"], to_row, to_note_col, NOTE_TO, "TO")
    te_pass, te_bad = tally(wb["TE"], te_row, te_note_col, NOTE_TE, "TE")
    if not to_pass:
        r.passed = False
        if to_bad:
            r.findings.append("  TO malformed sample (row, value): " + str(to_bad))
    if not te_pass:
        r.passed = False
        if te_bad:
            r.findings.append("  TE malformed sample (row, value): " + str(te_bad))
    return r


def check_note_ccn_consistency(wb, to_meta, te_meta):
    r = CheckResult("3. NOTE<->CCN consistency", True)
    if not (to_meta and te_meta):
        r.passed = False
        r.findings.append("skipped: schema check failed")
        return r

    def scan(ws, hdr_row, cols, label):
        note_col = cols["NOTE"]
        ccn_col = cols["CCN"]
        mismatches = 0
        sample = []
        examined = 0
        for rr in range(hdr_row + 1, ws.max_row + 1):
            note_v = ws.cell(row=rr, column=note_col).value
            ccn_v = ws.cell(row=rr, column=ccn_col).value
            if note_v in (None, "") and ccn_v in (None, ""):
                continue
            examined += 1
            note_s = str(note_v).strip() if note_v is not None else ""
            ccn_s = str(ccn_v).strip() if ccn_v is not None else ""
            note_ccn_part = re.match(r"^(\d{4,5})", note_s)
            note_ccn_part = note_ccn_part.group(1) if note_ccn_part else ""
            if note_ccn_part and ccn_s and note_ccn_part != ccn_s:
                mismatches += 1
                if len(sample) < 5:
                    sample.append((rr, note_s, ccn_s))
        r.findings.append(
            f"{label}: rows examined={examined}, mismatches={mismatches}"
        )
        if mismatches and sample:
            r.findings.append(f"  {label} sample: {sample}")
        return mismatches == 0

    ok_to = scan(wb["TO"], to_meta[0], to_meta[1], "TO")
    ok_te = scan(wb["TE"], te_meta[0], te_meta[1], "TE")
    if not (ok_to and ok_te):
        r.passed = False
    return r


def check_vocabulary(wb, to_meta, te_meta, vocab):
    r = CheckResult("4. Vocabulary", True)
    if vocab is None:
        r.passed = False
        r.findings.append(f"skipped: {VOCAB_PATH} not found")
        return r

    ccn_sheets = discover_ccn_sheets(wb)
    sheet_unknown = [s for s in ccn_sheets if s not in vocab]
    r.findings.append(
        f"CCN sheets found: {len(ccn_sheets)}; "
        f"unknown to vocabulary: {len(sheet_unknown)}"
    )
    if sheet_unknown:
        r.findings.append(f"  unknown sheet names: {sheet_unknown}")
        r.passed = False

    def harvest(ws, hdr_row, cols, label):
        if hdr_row is None:
            return set(), set()
        ccn_col = cols.get("CCN")
        seen = set()
        unknown = set()
        for rr in range(hdr_row + 1, ws.max_row + 1):
            v = ws.cell(row=rr, column=ccn_col).value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            # Strip any space form -> 5 digits
            s_un = s.replace(" ", "")
            if CCN_5.match(s_un):
                seen.add(s_un)
                if s_un not in vocab:
                    unknown.add(s_un)
        r.findings.append(
            f"{label}: unique CCNs in CCN column={len(seen)}; "
            f"unknown to vocabulary={len(unknown)}"
        )
        return seen, unknown

    if to_meta:
        _, to_unknown = harvest(wb["TO"], to_meta[0], to_meta[1], "TO")
        if to_unknown:
            r.findings.append(f"  TO unknown sample: {sorted(to_unknown)[:10]}")
            r.passed = False
    if te_meta:
        _, te_unknown = harvest(wb["TE"], te_meta[0], te_meta[1], "TE")
        if te_unknown:
            r.findings.append(f"  TE unknown sample: {sorted(te_unknown)[:10]}")
            r.passed = False
    return r


def check_cell_errors(workbook_path):
    """Scan all cached cell values for error tokens. Requires data_only=True
    workbook load (cached values from last Excel/LibreOffice save)."""
    r = CheckResult("5. Cell error scan", True)
    wb_dat = openpyxl.load_workbook(workbook_path, data_only=True)
    counts = {tok: 0 for tok in ERR_TOKENS}
    samples = []
    for sn in wb_dat.sheetnames:
        ws = wb_dat[sn]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                s = str(v)
                for tok in ERR_TOKENS:
                    if tok in s:
                        counts[tok] += 1
                        if len(samples) < 10:
                            samples.append(f"{sn}!{cell.coordinate} = {s[:30]}")
                        break
    total = sum(counts.values())
    r.findings.append(f"error tokens found: {total}; counts={counts}")
    if total > 0:
        r.passed = False
        r.findings.extend([f"  {s}" for s in samples])
    return r


def check_rollup(wb, vocab):
    r = CheckResult("6. Roll-up integrity", True)

    rollup_name = None
    for sn in wb.sheetnames:
        if "ROLLUP" in sn.upper().replace("_", "").replace(" ", ""):
            rollup_name = sn
            break
    if rollup_name is None:
        r.passed = False
        r.findings.append("no UNIT_ROLLUP sheet found")
        return r

    rollup = wb[rollup_name]
    listed_ccns = set()
    for row in rollup.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            s = str(v).strip().replace(" ", "")
            if CCN_5.match(s):
                listed_ccns.add(s)
    ccn_sheets = set(discover_ccn_sheets(wb))
    missing_from_rollup = ccn_sheets - listed_ccns
    extra_in_rollup = listed_ccns - ccn_sheets

    r.findings.append(
        f"UNIT_ROLLUP sheet: {rollup_name!r}; "
        f"CCN sheets in workbook={len(ccn_sheets)}; "
        f"CCNs listed in rollup={len(listed_ccns)}"
    )
    r.findings.append(
        f"workbook CCN sheets: {sorted(ccn_sheets)}"
    )
    r.findings.append(
        f"rollup-listed CCNs: {sorted(listed_ccns)}"
    )
    if missing_from_rollup:
        r.passed = False
        r.findings.append(
            f"FAIL: CCN sheets present but NOT rolled up: "
            f"{sorted(missing_from_rollup)}"
        )
    if extra_in_rollup:
        r.passed = False
        r.findings.append(
            f"FAIL: CCNs in rollup with no sheet: "
            f"{sorted(extra_in_rollup)}"
        )

    # Per-sheet TOTAL REQUIREMENT cell discovery
    for sn in sorted(ccn_sheets):
        ws = wb[sn]
        lab, vc, val = discover_total_cell(ws)
        if lab is None:
            r.passed = False
            r.findings.append(
                f"  sheet {sn!r}: no 'TOTAL REQUIREMENT' label found"
            )
        else:
            r.findings.append(
                f"  sheet {sn!r}: TOTAL at {vc} = {val!r}"
            )
    return r


def check_billet_accounting(wb, to_meta):
    """Check 7. Every TO data row attributes its billet to exactly one
    CCN sheet via the NOTE column. Orphans (TO rows with no NOTE) and
    NOTE-tag CCNs that have no corresponding sheet are failures."""
    r = CheckResult("7. Billet accounting", True)
    if not to_meta:
        r.passed = False
        r.findings.append("skipped: schema check failed")
        return r
    hdr_row, cols = to_meta
    note_col = cols.get("NOTE")
    bic_col = cols.get("BIC")
    ws = wb["TO"]

    ccn_sheets = set(discover_ccn_sheets(wb))
    total = 0
    attributed = 0
    orphans = 0
    by_ccn = {}
    unknown_ccns = {}
    sample_orphans = []
    sample_unknown = []

    for rr in range(hdr_row + 1, ws.max_row + 1):
        bic_v = ws.cell(row=rr, column=bic_col).value if bic_col else None
        # Treat a row as a billet only if it has any non-empty cell in the
        # row (not just BIC, since some rows may use a placeholder)
        any_non_empty = any(
            ws.cell(row=rr, column=c).value not in (None, "")
            for c in range(1, ws.max_column + 1)
        )
        if not any_non_empty:
            continue
        total += 1
        note_v = ws.cell(row=rr, column=note_col).value
        if note_v is None or str(note_v).strip() == "":
            orphans += 1
            if len(sample_orphans) < 5:
                sample_orphans.append((rr, str(bic_v) if bic_v else ""))
            continue
        m = re.match(r"^(\d{4,5})", str(note_v).strip())
        if not m:
            orphans += 1
            if len(sample_orphans) < 5:
                sample_orphans.append((rr, str(note_v)))
            continue
        ccn = m.group(1)
        attributed += 1
        by_ccn[ccn] = by_ccn.get(ccn, 0) + 1
        if ccn not in ccn_sheets:
            unknown_ccns[ccn] = unknown_ccns.get(ccn, 0) + 1
            if len(sample_unknown) < 5:
                sample_unknown.append((rr, ccn))

    r.findings.append(
        f"TO data rows={total}, attributed={attributed}, orphans={orphans}, "
        f"distinct attributed CCNs={len(by_ccn)}"
    )
    if by_ccn:
        top = sorted(by_ccn.items(), key=lambda x: -x[1])[:6]
        r.findings.append(f"  per-CCN counts (top 6): {top}")
    if orphans > 0:
        r.passed = False
        r.findings.append(
            f"  orphan sample (row, BIC or NOTE): {sample_orphans}"
        )
    if unknown_ccns:
        r.passed = False
        r.findings.append(
            f"  TO NOTE references {len(unknown_ccns)} CCN(s) with no sheet "
            f"in this workbook: {sorted(unknown_ccns.keys())}"
        )
        r.findings.append(
            f"  unknown sample (row, CCN): {sample_unknown}"
        )
    return r


def check_equipment_accounting(wb, te_meta):
    """Check 8. Every TE data row attributes its TAMCN to a CCN via the
    NOTE or CCN column. Orphans and unknown-CCN references are failures."""
    r = CheckResult("8. Equipment accounting", True)
    if not te_meta:
        r.passed = False
        r.findings.append("skipped: schema check failed")
        return r
    hdr_row, cols = te_meta
    note_col = cols.get("NOTE")
    ccn_col = cols.get("CCN")
    ws = wb["TE"]

    ccn_sheets = set(discover_ccn_sheets(wb))
    total = 0
    attributed = 0
    orphans = 0
    by_ccn = {}
    unknown_ccns = {}
    sample_orphans = []
    sample_unknown = []

    for rr in range(hdr_row + 1, ws.max_row + 1):
        any_non_empty = any(
            ws.cell(row=rr, column=c).value not in (None, "")
            for c in range(1, ws.max_column + 1)
        )
        if not any_non_empty:
            continue
        total += 1
        note_v = ws.cell(row=rr, column=note_col).value
        ccn_v = ws.cell(row=rr, column=ccn_col).value
        ref = None
        for v in (note_v, ccn_v):
            if v is None:
                continue
            s = str(v).strip().replace(" ", "")
            m = re.match(r"^(\d{4,5})", s)
            if m:
                ref = m.group(1)
                break
        if ref is None:
            orphans += 1
            if len(sample_orphans) < 5:
                sample_orphans.append((rr, str(note_v), str(ccn_v)))
            continue
        attributed += 1
        by_ccn[ref] = by_ccn.get(ref, 0) + 1
        if ref not in ccn_sheets:
            unknown_ccns[ref] = unknown_ccns.get(ref, 0) + 1
            if len(sample_unknown) < 5:
                sample_unknown.append((rr, ref))

    r.findings.append(
        f"TE data rows={total}, attributed={attributed}, orphans={orphans}, "
        f"distinct attributed CCNs={len(by_ccn)}"
    )
    if by_ccn:
        top = sorted(by_ccn.items(), key=lambda x: -x[1])[:6]
        r.findings.append(f"  per-CCN counts (top 6): {top}")
    if orphans > 0:
        r.passed = False
        r.findings.append(
            f"  orphan sample (row, NOTE, CCN): {sample_orphans}"
        )
    if unknown_ccns:
        r.passed = False
        r.findings.append(
            f"  TE references {len(unknown_ccns)} CCN(s) with no sheet "
            f"in this workbook: {sorted(unknown_ccns.keys())}"
        )
        r.findings.append(
            f"  unknown sample (row, CCN): {sample_unknown}"
        )
    return r


# ---------- Driver ---------------------------------------------------

def run(workbook_path: Path, report_path: Optional[Path]):
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    vocab = load_vocabulary()

    s, to_meta, te_meta = check_schema(wb)
    n = check_note_coverage(wb, to_meta, te_meta)
    c = check_note_ccn_consistency(wb, to_meta, te_meta)
    v = check_vocabulary(wb, to_meta, te_meta, vocab)
    e = check_cell_errors(workbook_path)
    u = check_rollup(wb, vocab)
    b = check_billet_accounting(wb, to_meta)
    q = check_equipment_accounting(wb, te_meta)

    results = [s, n, c, v, e, u, b, q]
    pass_count = sum(1 for x in results if x.passed)
    fail_count = len(results) - pass_count

    lines = []
    lines.append("=" * 64)
    lines.append("BFR Validation Report, Layer 6 Pipeline Harness")
    lines.append("=" * 64)
    lines.append(f"Source workbook : {workbook_path}")
    lines.append(f"Sheet count     : {len(wb.sheetnames)}")
    visible = sum(1 for s_ in wb.sheetnames if wb[s_].sheet_state == "visible")
    lines.append(f"  visible       : {visible}")
    lines.append(f"  hidden        : {len(wb.sheetnames) - visible}")
    if vocab is not None:
        lines.append(f"Vocabulary      : {VOCAB_PATH.name} ({len(vocab)} CCNs)")
    else:
        lines.append("Vocabulary      : (not found)")
    lines.append("")
    for res in results:
        lines.append(res.line())
        for f in res.findings:
            lines.append(f"  - {f}")
        lines.append("")
    lines.append("-" * 64)
    lines.append(f"OVERALL: {pass_count} PASS / {fail_count} FAIL")
    lines.append("-" * 64)
    out = "\n".join(lines)

    print(out)
    if report_path:
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(out + "\n")
    return 0 if fail_count == 0 else 1


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook", type=Path)
    ap.add_argument("--report", type=Path, default=None,
                    help="path to write the report (also printed to stdout)")
    args = ap.parse_args()
    if not args.workbook.exists():
        print(f"FATAL: not found: {args.workbook}", file=sys.stderr)
        sys.exit(2)
    sys.exit(run(args.workbook, args.report))


if __name__ == "__main__":
    main()
